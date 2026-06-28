#!/usr/bin/env python3
"""Moteur de génération des classeurs Excel de suivi des loyers (part locataire + part CAF).

Un fichier .xlsx par bailleur. Le contenu (colonnes, feuilles, documents) est piloté par les
`modules` activés.

Organisation du classeur :
  * Guide                 : mode d'emploi.
  * Locataires            : référentiel (nom, type/identifiant/adresse du bien, loyers, dépôt).
  * Une feuille PAR locataire : saisie mensuelle, nommée par le n° d'appartement (ou le nom de
    la maison). Lisible même avec beaucoup de locataires.
  * Données (masquée)     : consolide toutes les feuilles locataire par formules ; alimente le
    Bilan et les documents (SUMIFS sur plages nommées). Aucune double saisie.
  * Bilan                 : synthèse par locataire.
  * Quittance / Avis d'échéance / Lettre de relance : documents à imprimer (si module actif).

Particularités :
  * Les lignes ne couvrent que la période d'activité de chaque locataire (entrée -> sortie) :
    gère les rotations fréquentes.
  * Régénération sur un fichier existant : les montants déjà saisis (CAF reçue, part locataire,
    dates) sont préservés (clé = locataire + année + mois).

Usage CLI / Docker :  python generer_suivi_loyers.py <config.yaml> [dossier_sortie]
"""

from __future__ import annotations

import sys
import re
import math
import shutil
import calendar
import warnings
import datetime as dt
from copy import copy
from dataclasses import dataclass
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties, Font as PoliceDessin,
)

# Chaîne numérique simple (autorise la virgule décimale d'une config éditée à la main).
_NUM_TXT = re.compile(r"^-?\d+(?:[.,]\d+)?$")

# --------------------------------------------------------------------------- #
# Constantes
# --------------------------------------------------------------------------- #

MOIS = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

FMT_EURO = '#,##0.00\\ "€"'
FMT_DATE = "DD/MM/YYYY"
FMT_PCT = "0.0%"

# --- Identité graphique : registre de thèmes sélectionnables --------------- #
# Un thème = une entrée du registre (couleurs + onglets par rôle). La clé config
# `theme` choisit l'entrée ; absente ou inconnue, on retombe sur le défaut
# (« classique » = look historique). Tout le rendu lit l'objet `CHARTE` résolu au
# début de generer_workbook, jamais des couleurs en dur : ajouter un thème ne
# touche pas au cœur (extensibilité sans modification du cœur).

POLICE_DEFAUT = "Tahoma"   # police d'origine sur tout Windows (Aptos non stock OS)

THEMES: dict[str, dict] = {
    "classique": {
        "label": "Classique",
        "teinte": "bleu marine",
        "fond": "FFFFFF",
        "primaire": "1F4E79",   # bandeaux, titres, onglets « système »
        "saisie": "FFF7E6",     # jaune pâle : cellules à remplir
        "calc": "EEF3F8",       # bleu très pâle : cellules calculées
        "solde": "C6EFCE",
        "trop": "FFEB9C",
        "partiel": "FFC7CE",
        "attente": "E7E6E6",
        "lien": "0563C1",       # bleu hyperlien
        "locat": "548235",      # onglets feuilles locataire
        "docs": "C55A11",       # onglets documents
        "donnees": "808080",    # onglet Données (masqué)
        "table": "TableStyleMedium2",  # style du tableau Locataires (famille de teinte)
    },
    "refresh-bleu": {
        "label": "Refresh bleu",
        "teinte": "bleu acier",
        "fond": "F0EEE6",
        "primaire": "234B73",
        "saisie": "FBF3DC",
        "calc": "E9EFF5",
        "solde": "C6EFCE",
        "trop": "FFEB9C",
        "partiel": "FFC7CE",
        "attente": "E7E6E6",
        "lien": "3B6FB0",
        "locat": "4E7A4E",
        "docs": "C06A3E",
        "donnees": "8C8780",
        "table": "TableStyleMedium2",   # bleu
    },
    "denim-rust": {
        "label": "Denim & Rust",
        "teinte": "bleu denim & rouille",
        "fond": "F0EEE6",
        "primaire": "33455C",
        "saisie": "FBF2DC",
        "calc": "E9EDF1",
        "solde": "C9E2C4",
        "trop": "F6E2B3",
        "partiel": "F1C9C2",
        "attente": "E6E3DD",
        "lien": "3B6FB0",
        "locat": "5E806B",
        "docs": "B5603F",
        "donnees": "8C8780",
        "table": "TableStyleMedium3",   # rouille
    },
    "monochrome-noir": {
        "label": "Monochrome Noir",
        "teinte": "noir & gris",
        "fond": "F5F5F4",
        "primaire": "141414",
        "saisie": "F0EFEA",
        "calc": "EEEEED",
        "solde": "C9E2C4",
        "trop": "F6E2B3",
        "partiel": "F1C9C2",
        "attente": "E6E3DD",
        "lien": "444444",
        "locat": "3D3D3D",
        "docs": "6E6E6E",
        "donnees": "9A9A9A",
        "table": "TableStyleMedium4",   # gris neutre
    },
    "ocean-deep": {
        "label": "Ocean Deep",
        "teinte": "bleu océan",
        "fond": "EEF2F4",
        "primaire": "0B3954",
        "saisie": "F6EFDD",
        "calc": "E4EDF0",
        "solde": "C9E2C4",
        "trop": "F6E2B3",
        "partiel": "F1C9C2",
        "attente": "E6E3DD",
        "lien": "1E6F8C",
        "locat": "2A9D8F",
        "docs": "E76F51",
        "donnees": "87A0A8",
        "table": "TableStyleMedium6",   # bleu clair (proche turquoise)
    },
}
THEME_DEFAUT = "classique"


@dataclass(frozen=True)
class Charte:
    """Couleurs résolues d'un thème, lues par tout le rendu (aucune couleur en dur)."""
    primaire: str
    entete_txt: str
    saisie: str
    calc: str
    solde: str
    trop: str
    partiel: str
    attente: str
    lien: str
    onglet_systeme: str
    onglet_locataire: str
    onglet_document: str
    onglet_donnees: str
    fond: str
    police: str
    table: str


def resoudre_charte(theme: str | None = None, police: str | None = None) -> Charte:
    """Charte d'un thème. Thème absent → défaut ; inconnu → défaut + avertissement."""
    if theme and theme not in THEMES:
        warnings.warn(
            f"Thème « {theme} » inconnu : utilisation du thème par défaut "
            f"« {THEME_DEFAUT} ». Thèmes disponibles : {', '.join(sorted(THEMES))}.",
            stacklevel=2)
        theme = None
    t = THEMES[theme or THEME_DEFAUT]
    return Charte(
        primaire=t["primaire"],
        entete_txt="FFFFFF",            # texte blanc sur bandeau, tous thèmes
        saisie=t["saisie"],
        calc=t["calc"],
        solde=t["solde"],
        trop=t["trop"],
        partiel=t["partiel"],
        attente=t["attente"],
        lien=t["lien"],
        onglet_systeme=t["primaire"],
        onglet_locataire=t["locat"],
        onglet_document=t["docs"],
        onglet_donnees=t["donnees"],
        fond=t["fond"],
        police=police or POLICE_DEFAUT,
        table=t.get("table", "TableStyleMedium2"),
    )


# Charte active. Défaut (look classique) au niveau module pour tout appel direct ;
# generer_workbook la réassigne selon la config avant de construire les feuilles.
CHARTE = resoudre_charte()

# Échelle de titres unifiée et hauteur de la ligne d'en-tête (en-têtes sur 2 lignes).
TITRE_H1 = 16
TITRE_H2 = 12
HAUTEUR_ENTETE = 30

# Valeurs officielles de l'IRL (série trimestrielle INSEE), à recopier dans l'onglet IRL.
URL_IRL_INSEE = "https://www.insee.fr/fr/statistiques/serie/001515333"

_THIN = Side(style="thin", color="BFBFBF")
BORDURE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

TYPES_BIEN = ["Appartement", "Maison"]

# Colonnes de saisie utilisateur à préserver lors d'une régénération.
COLS_SAISIE = ("caf_recu", "caf_date", "loc_recu", "loc_date")

# Feuilles « système » (tout le reste = une feuille locataire).
FEUILLES_SYSTEME = {"Guide", "Locataires", "Données", "Bilan", "Tableau de bord",
                    "Régularisation charges", "Révision IRL",
                    "Quittance", "Avis d'échéance", "Lettre de relance"}

TRIMESTRES = ["T1", "T2", "T3", "T4"]

# Motifs de départ proposés (liste déroulante, saisie libre possible).
OBSERVATIONS = [
    "Fin de bail", "Congé du locataire", "Congé du bailleur (vente)",
    "Congé du bailleur (reprise)", "Loyer impayé", "Travaux",
    "Dégradations (retenue sur caution)", "Départ anticipé", "Abandon de logement",
]

MODULES_DEFAUT = {
    "mode_charges": "comprises",     # "comprises" | "separees" | "sans"
    "loyer_nu_charges": True,        # déprécié (rétro-compat) : mappé vers mode_charges
    "caf": True,
    "depot_garantie": True,
    "documents": True,               # quittance + avis d'échéance + lettre de relance
    "tableau_bord": True,            # onglet de graphiques
    "irl": False,
    "regularisation_charges": False,
}

MODES_CHARGES = ["comprises", "separees", "sans"]


def _flags_charges(cfg: dict):
    """(a_charges, charges_separees, mode) à partir du mode de charges du bailleur."""
    m = cfg["modules"]
    mode = m.get("mode_charges")
    if mode not in MODES_CHARGES:
        mode = "separees" if m.get("loyer_nu_charges", True) else "sans"
    return mode in ("comprises", "separees"), mode == "separees", mode

# Version du schéma de config. Incrémenter quand un changement nécessite une migration ;
# migrer_config() doit alors gérer la transition depuis les versions antérieures.
CONFIG_VERSION = 1


# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #

def migrer_config(raw: dict) -> tuple[dict, list[str]]:
    """Normalise une config, y compris ancienne, de façon TOLÉRANTE.

    Ne lève pas pour des champs manquants : complète au mieux et renvoie la liste des
    adaptations effectuées (pour informer l'utilisateur). Sert de point unique de
    rétro-compatibilité quand le schéma évolue (cf. CONFIG_VERSION).
    """
    if not isinstance(raw, dict):
        raise ValueError("Fichier de configuration illisible (format inattendu).")

    avertis: list[str] = []
    cfg = dict(raw)
    ver = cfg.get("version")
    if isinstance(ver, int) and ver > CONFIG_VERSION:
        avertis.append(
            f"Configuration créée par une version plus récente (v{ver}) que ce logiciel "
            f"(v{CONFIG_VERSION}) : des options récentes peuvent être ignorées.")
    elif ver not in (None, CONFIG_VERSION):
        avertis.append(f"Configuration version {ver} adaptée à la version {CONFIG_VERSION}.")

    cfg["bailleur"] = dict(cfg.get("bailleur") or {})
    cfg["periode"] = dict(cfg.get("periode") or {})

    modules = {}
    for k, v in (cfg.get("modules") or {}).items():
        kl = str(k).strip().lower()
        if kl == "quittances":
            kl = "documents"
            avertis.append("Ancien module « quittances » converti en « documents ».")
        if kl not in MODULES_DEFAUT:
            avertis.append(f"Module inconnu ignoré : « {k} ».")
            continue
        modules[kl] = v
    # Rétro-compat : ancien booléen loyer_nu_charges -> mode_charges.
    if "mode_charges" not in modules and "loyer_nu_charges" in modules:
        modules["mode_charges"] = "separees" if modules["loyer_nu_charges"] else "sans"
    cfg["modules"] = modules

    locataires = []
    converti_bien = False
    for loc in (cfg.get("locataires") or []):
        if not isinstance(loc, dict):
            avertis.append("Un locataire au format inattendu a été ignoré.")
            continue
        loc = dict(loc)
        # Coercition douce : « 100,5 » -> « 100.5 » pour les montants saisis à la main.
        for champ in ("loyer_nu", "charges", "loyer", "loyer_total", "part_caf", "depot_garantie"):
            v = loc.get(champ)
            if isinstance(v, str) and _NUM_TXT.match(v.strip()):
                loc[champ] = v.strip().replace(",", ".")
        loc.setdefault("type_bien", TYPES_BIEN[0])
        if not loc.get("identifiant"):
            loc["identifiant"] = loc.get("bien") or loc.get("nom") or ""
            if loc.get("bien"):
                converti_bien = True
        locataires.append(loc)
    if converti_bien:
        avertis.append("Ancien champ « bien » repris comme identifiant du logement.")
    cfg["locataires"] = locataires

    # Apparence : thème (couleurs) + police, tolérants. Thème inconnu -> défaut + avertissement.
    theme = cfg.get("theme")
    if theme is not None:
        theme = str(theme).strip().lower()
        if theme and theme not in THEMES:
            avertis.append(
                f"Thème « {cfg.get('theme')} » inconnu : thème par défaut "
                f"« {THEME_DEFAUT} » utilisé.")
            theme = THEME_DEFAUT
        cfg["theme"] = theme or THEME_DEFAUT
    police = cfg.get("police")
    if police is not None:
        police = str(police).strip()
        cfg["police"] = police or POLICE_DEFAUT

    cfg["version"] = CONFIG_VERSION
    return cfg, avertis


def valider_config(raw: dict) -> dict:
    cfg, _ = migrer_config(raw)

    bailleur = cfg["bailleur"]
    if not bailleur.get("nom"):
        raise ValueError("Le nom du bailleur est obligatoire.")

    periode = cfg["periode"]
    annee_debut = int(periode.get("annee_debut", dt.date.today().year))
    annee_fin = int(periode.get("annee_fin", annee_debut))
    if annee_fin < annee_debut:
        raise ValueError("L'année de fin doit être supérieure ou égale à l'année de début.")

    modules = dict(MODULES_DEFAUT)
    modules.update(cfg["modules"])

    locataires = cfg["locataires"]
    if not locataires:
        raise ValueError("Il faut au moins un locataire.")

    vus: dict[str, int] = {}
    for i, loc in enumerate(locataires, 1):
        if not loc.get("nom"):
            raise ValueError(f"Le locataire #{i} n'a pas de nom.")
        ident = _identite(loc)
        for champ in ("loyer_nu", "charges", "loyer", "loyer_total", "part_caf",
                      "depot_garantie"):
            if loc.get(champ) not in (None, ""):
                try:
                    x = float(loc[champ])
                except (TypeError, ValueError):
                    raise ValueError(
                        f"Locataire « {ident} » : {champ} non numérique ({loc[champ]!r}).") from None
                if not math.isfinite(x):
                    raise ValueError(f"Locataire « {ident} » : {champ} invalide.")
        try:
            de = _date(loc.get("date_entree"))
        except ValueError:
            raise ValueError(
                f"Locataire « {ident} » : date d'entrée invalide ({loc.get('date_entree')!r}).") from None
        try:
            ds = _date(loc.get("date_sortie"))
        except ValueError:
            raise ValueError(
                f"Locataire « {ident} » : date de sortie invalide ({loc.get('date_sortie')!r}).") from None
        if de and ds and ds < de:
            raise ValueError(f"Locataire « {ident} » : date de sortie avant la date d'entrée.")
        if ident in vus:
            raise ValueError(
                f"Deux locataires portent la même identité « {ident} ». "
                "Ajoutez un prénom ou différenciez-les (les saisies seraient mélangées).")
        vus[ident] = i

    return {
        "bailleur": bailleur,
        "annee_debut": annee_debut,
        "annee_fin": annee_fin,
        "modules": modules,
        "locataires": locataires,
        "demo": bool(raw.get("demo")),   # pré-remplissage de démonstration (exemples uniquement)
        "theme": cfg.get("theme") or THEME_DEFAUT,
        "police": cfg.get("police") or POLICE_DEFAUT,
    }


def charger_config(chemin: Path) -> dict:
    with Path(chemin).open(encoding="utf-8") as f:
        return valider_config(yaml.safe_load(f))


def _num(loc: dict, *cles) -> float | None:
    for c in cles:
        if loc.get(c) is not None:
            v = float(loc[c])
            if not math.isfinite(v):
                raise ValueError(f"Montant invalide : {loc[c]!r}")
            return v
    return None


def _date(val) -> dt.date | None:
    if val is None or val == "":
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    return dt.date.fromisoformat(str(val))


def _mois_actifs(loc: dict, annee_debut: int, annee_fin: int) -> list[tuple[int, int]]:
    de = _date(loc.get("date_entree"))
    ds = _date(loc.get("date_sortie"))
    debut = (de.year, de.month) if de else (annee_debut, 1)
    fin = (ds.year, ds.month) if ds else (annee_fin, 12)
    return [(a, m) for a in range(annee_debut, annee_fin + 1)
            for m in range(1, 13) if debut <= (a, m) <= fin]


def _prorata_suffixe(loc: dict, annee: int, mois: int) -> str:
    """Suffixe de formule « *jours/jours_du_mois » pour un mois d'entrée/sortie partiel.

    Vide si mois plein. Jours réels du mois (gère février et les mois de 31 jours).
    """
    de, ds = _date(loc.get("date_entree")), _date(loc.get("date_sortie"))
    nb = calendar.monthrange(annee, mois)[1]
    premier, dernier = 1, nb
    if de and (de.year, de.month) == (annee, mois):
        premier = de.day
    if ds and (ds.year, ds.month) == (annee, mois):
        dernier = ds.day
    jours = dernier - premier + 1
    if jours <= 0 or jours >= nb:
        return ""
    return f"*{jours}/{nb}"


def _prorata_facteur(loc: dict, annee: int, mois: int) -> float:
    """Fraction du mois réellement occupée (1.0 si mois plein). Pendant du suffixe ci-dessus."""
    de, ds = _date(loc.get("date_entree")), _date(loc.get("date_sortie"))
    nb = calendar.monthrange(annee, mois)[1]
    premier, dernier = 1, nb
    if de and (de.year, de.month) == (annee, mois):
        premier = de.day
    if ds and (ds.year, ds.month) == (annee, mois):
        dernier = ds.day
    return max(0, dernier - premier + 1) / nb


def _trimestre_de(d) -> str:
    """Trimestre civil (« T1 ».. « T4 ») d'une date ; « T1 » par défaut si absente."""
    date = _date(d)
    return f"T{(date.month - 1) // 3 + 1}" if date else "T1"


def _annees_actives(loc: dict, annee_debut: int, annee_fin: int) -> list[int]:
    """Années pendant lesquelles le locataire est présent (au moins un mois actif)."""
    return sorted({a for (a, _m) in _mois_actifs(loc, annee_debut, annee_fin)})


# Mois <= ce repère : paiement de démonstration saisi ; après : laissé vide (« À encaisser »).
# Fixe (pas la date du jour) pour que les classeurs d'exemple soient reproductibles.
DEMO_CUTOFF = (2026, 3)


def _saisies_demo(cfg: dict) -> dict:
    """Saisies de démonstration pour les exemples : loyers reçus jusqu'à DEMO_CUTOFF,
    soldés pour tous, avec un impayé thématique en fin de bail (observation « impayé »).
    Déterministe. Format identique à recolter_saisies (clé (identité, année, mois))."""
    split, _csep, _mode = _flags_charges(cfg)
    caf = cfg["modules"]["caf"]
    saisies: dict = {}
    for loc in cfg["locataires"]:
        ident = _identite(loc)
        if split:
            base = (_num(loc, "loyer_nu") or 0) + (_num(loc, "charges") or 0)
        else:
            base = _num(loc, "loyer", "loyer_total", "loyer_nu") or 0
        part_caf = (_num(loc, "part_caf") or 0) if caf else 0
        mois = [am for am in _mois_actifs(loc, cfg["annee_debut"], cfg["annee_fin"])
                if am <= DEMO_CUTOFF]
        impaye = "impayé" in str(loc.get("observation") or "").lower()
        fin_impaye = set(mois[-2:]) if impaye else set()
        for (annee, m) in mois:
            du_total = round(base * _prorata_facteur(loc, annee, m), 2)
            caf_recu = round(part_caf, 2)            # la CAF n'est pas proratisée
            loc_du = max(0.0, round(du_total - caf_recu, 2))
            loc_recu = round(loc_du * 0.4, 2) if (annee, m) in fin_impaye else loc_du
            entree = {"loc_recu": loc_recu, "loc_date": dt.date(annee, m, 5)}
            if caf and caf_recu:
                entree.update(caf_recu=caf_recu, caf_date=dt.date(annee, m, 5))
            saisies[(ident, int(annee), MOIS[m - 1])] = entree
    return saisies


def _slug(nom: str) -> str:
    s = re.sub(r"[^\w\-]+", "_", nom.strip(), flags=re.UNICODE)
    return s.strip("_") or "bailleur"


def base_fichier(bailleur: dict) -> str:
    """Base du nom de fichier : nom de la SCI (si renseigné) + nom du bailleur."""
    nom = str(bailleur.get("nom") or "").strip()
    prenom = str(bailleur.get("prenom") or "").strip()
    perso = f"{nom} {prenom}".strip() if prenom else nom
    if bailleur.get("sci") and str(bailleur.get("sci_nom") or "").strip():
        return f"{bailleur['sci_nom']} {perso}".strip()
    return perso or nom


def base_slug(bailleur: dict) -> str:
    """Base assainie pour nom de fichier (contrat public pour l'interface)."""
    return _slug(base_fichier(bailleur))


def _identite(loc: dict) -> str:
    """Identité d'un locataire (clé interne) = « Nom Prénom » (ou nom seul si pas de prénom)."""
    nom = str(loc.get("nom") or "").strip()
    prenom = str(loc.get("prenom") or "").strip()
    return f"{nom} {prenom}".strip() if prenom else nom


_CAR_INTERDITS = set('[]:*?/\\')


def _nom_feuille(ident: str, pris: set) -> str:
    """Nom de feuille Excel valide et unique (<= 31 car., sans caractères interdits)."""
    base = "".join(c for c in str(ident) if c not in _CAR_INTERDITS).strip()[:31] or "Locataire"
    nom, i = base, 2
    while nom in pris:
        suff = f" ({i})"
        nom = base[:31 - len(suff)] + suff
        i += 1
    pris.add(nom)
    return nom


def _ref(feuille: str, cellule: str) -> str:
    """Référence inter-feuilles robuste (gère espaces et apostrophes)."""
    return "'" + feuille.replace("'", "''") + "'!" + cellule


# --------------------------------------------------------------------------- #
# Styles
# --------------------------------------------------------------------------- #

def style_entete(cell) -> None:
    cell.font = Font(bold=True, color=CHARTE.entete_txt)
    cell.fill = PatternFill("solid", fgColor=CHARTE.primaire)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDURE


def style_titre(cell, niveau: int = TITRE_H1):
    """Titre d'onglet (TITRE_H1) ou de section (TITRE_H2), couleur d'identité."""
    cell.font = Font(bold=True, size=niveau, color=CHARTE.primaire)
    return cell


def regler_hauteur_entete(ws, row: int = 1) -> None:
    """Hauteur de charte pour la ligne d'en-tête (en-têtes sur 2 lignes lisibles)."""
    ws.row_dimensions[row].height = HAUTEUR_ENTETE


def _fill_cf(couleur: str) -> PatternFill:
    """Remplissage pour la mise en forme conditionnelle.

    Dans un format conditionnel (dxf), Excel lit la couleur sur bgColor, pas fgColor :
    il faut donc renseigner start_color ET end_color, sinon la couleur n'apparaît pas.
    """
    return PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")


# Préfixes qu'Excel interprète comme une formule. openpyxl écrit une chaîne commençant
# par « = » comme formule : on force le type texte pour neutraliser toute injection.
_PREFIXES_FORMULE = ("=", "+", "-", "@")


def _neutraliser(cell):
    """Force le type texte si la valeur (utilisateur) ressemble à une formule."""
    v = cell.value
    if isinstance(v, str) and v[:1] in _PREFIXES_FORMULE:
        cell.data_type = "s"
    return cell


def ecrire_texte(ws, row, column, valeur):
    """Écrit une valeur d'origine utilisateur sans risque d'interprétation en formule."""
    return _neutraliser(ws.cell(row, column, valeur))


def ecrire_lien(cell, texte: str, url: str):
    """Cellule hyperlien (texte cliquable, style lien)."""
    cell.value = texte
    cell.hyperlink = url
    cell.font = Font(color=CHARTE.lien, underline="single")
    return cell


def _formule_liste(valeurs) -> str:
    """Formule de validation « liste inline ». Excel plafonne à 255 car. : on échoue
    explicitement plutôt que de tronquer silencieusement (cf. ajout futur d'options)."""
    formule = '"%s"' % ",".join(str(v) for v in valeurs)
    if len(formule) > 255:
        raise ValueError(
            f"Liste de validation trop longue ({len(formule)} car. > 255). "
            "Basculer sur une plage nommée.")
    return formule


def style_cellule(cell, *, saisie=False, calc=False, fmt=None) -> None:
    if saisie:
        cell.fill = PatternFill("solid", fgColor=CHARTE.saisie)
    elif calc:
        cell.fill = PatternFill("solid", fgColor=CHARTE.calc)
    if fmt:
        cell.number_format = fmt
    cell.border = BORDURE


def appliquer_police(wb: Workbook, police: str) -> None:
    """Impose la police d'identité à toutes les cellules, en préservant les autres
    attributs (gras, taille, couleur, soulignement).

    Modifier le style « Normal » après écriture ne se propage pas (openpyxl fige le
    style à l'écriture) : on réécrit donc la police cellule par cellule en dernière
    passe. C'est le point unique d'application de la police du thème.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                f = cell.font
                if f is not None and f.name != police:
                    nf = copy(f)
                    nf.name = police
                    cell.font = nf
        for chart in getattr(ws, "_charts", []):
            _police_graphique(chart, police)


def _txpr(police: str, *, gras: bool = False) -> RichText:
    """Propriétés de texte (police) pour un élément de graphique (axe, légende)."""
    cp = CharacterProperties(latin=PoliceDessin(typeface=police), b=gras)
    return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])


def _police_graphique(chart, police: str) -> None:
    """Applique la police d'identité au texte d'un graphique (titre, axes, légende) :
    openpyxl ne couvre pas le texte des graphiques via les polices de cellule, il y
    resterait donc du Calibri."""
    chart.x_axis.txPr = _txpr(police)
    chart.y_axis.txPr = _txpr(police)
    if chart.legend is not None:
        chart.legend.txPr = _txpr(police)
    # Titre : texte déjà posé en RichText ; on force la police sur chaque paragraphe/run.
    rich = getattr(getattr(chart.title, "tx", None), "rich", None)
    if rich is not None:
        for para in rich.p:
            cp = CharacterProperties(latin=PoliceDessin(typeface=police), b=True)
            if para.pPr is None:
                para.pPr = ParagraphProperties()
            para.pPr.defRPr = cp
            for run in (para.r or []):
                run.rPr = cp


# Largeur de colonne = nombre de caractères de la police de base du classeur
# (fonts[0] = Calibri, figée par openpyxl et non modifiable proprement). Les largeurs
# sont calibrées pour Calibri ; une police plus large déborde. On élargit donc chaque
# colonne d'un facteur dépendant de la police (Calibri = 1.0 = référence). Un seul
# nombre à régler par police si le rendu reste trop juste.
FACTEUR_LARGEUR = {
    "Calibri": 1.0,
    "Tahoma": 1.10,
    "Verdana": 1.16,
    "Segoe UI": 1.04,
    "Arial": 1.06,
    "Georgia": 1.08,
    "Times New Roman": 1.0,
}


def ajuster_colonnes(wb: Workbook, police: str) -> None:
    """Compense la largeur des polices plus larges que Calibri (cf. FACTEUR_LARGEUR).

    N'agit que sur les largeurs explicitement posées : préserve les proportions
    calibrées à la main, ne touche pas aux colonnes laissées par défaut.
    """
    facteur = FACTEUR_LARGEUR.get(police, 1.0)
    if facteur == 1.0:
        return
    for ws in wb.worksheets:
        for dim in ws.column_dimensions.values():
            if dim.width:
                dim.width = round(dim.width * facteur, 1)


def mettre_en_page_impression(ws, zone: str | None = None, *, paysage: bool = False,
                              hauteur_pages: int = 1, centre: bool = True) -> None:
    """Prépare une feuille à l'impression : A4, ajusté à la page, marges sobres.

    - `zone` : plage d'impression (ex. « A1:E29 »). None = laisser le tableur déterminer
      la zone (utile pour les feuilles à graphiques, hors plage de cellules).
    - `hauteur_pages` : 1 = tout sur une page ; 0 = ajuste en largeur seulement et laisse
      déborder en hauteur sur plusieurs pages (feuilles larges, texte lisible).
    - N'altère PAS les largeurs de colonnes : fitToPage est un zoom appliqué à l'impression.

    fitToWidth/Height ne s'appliquent que si fitToPage est activé (piège openpyxl).
    Le quadrillage est déjà masqué par construire_*.
    """
    if zone:
        ws.print_area = zone
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if paysage else ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = hauteur_pages
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.9, bottom=0.8,
                                  header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = centre


# --------------------------------------------------------------------------- #
# Onglet Locataires (référentiel)
# --------------------------------------------------------------------------- #

def construire_locataires(wb: Workbook, cfg: dict) -> dict:
    mod = cfg["modules"]
    split, _csep, _mode = _flags_charges(cfg)
    caf, depot = mod["caf"], mod["depot_garantie"]

    ws = wb.create_sheet("Locataires")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHARTE.onglet_systeme

    # Colonne 1 = identité (clé pour VLOOKUP et listes déroulantes).
    cols: list[tuple[str, str]] = [
        ("locataire", "Locataire (Nom Prénom)"),
        ("identifiant", "N° appart. / Nom maison"),
        ("adresse", "Adresse du logement"),
        ("type_bien", "Type de bien"),
    ]
    if split:
        cols += [("loyer_nu", "Loyer nu (€)"), ("charges", "Charges (€)")]
    cols.append(("loyer_total", "Loyer total (€)"))
    if caf:
        cols.append(("part_caf", "Part CAF / APL (€)"))
    cols.append(("reste", "Reste à charge (€)"))
    if depot:
        cols.append(("depot", "Dépôt garantie (€)"))
    cols += [("date_entree", "Date entrée"), ("date_sortie", "Date sortie")]
    if depot:
        cols.append(("caution", "Caution rendue"))
    cols.append(("observation", "Observation (motif de départ)"))
    # Champs propres au bail, repris sur les documents (lus par VLOOKUP).
    docs = mod["documents"]
    if docs:
        cols += [("date_bail", "Date du bail"), ("mode_paiement", "Mode de paiement"),
                 ("jour_echeance", "Jour d'échéance")]

    idx = {champ: i + 1 for i, (champ, _) in enumerate(cols)}
    lettre = {champ: get_column_letter(i + 1) for i, (champ, _) in enumerate(cols)}

    for i, (_, titre) in enumerate(cols, 1):
        style_entete(ws.cell(row=1, column=i, value=titre))
    regler_hauteur_entete(ws, 1)

    largeurs = {"locataire": 22, "type_bien": 14, "identifiant": 20, "adresse": 28,
                "loyer_nu": 13, "charges": 13, "loyer_total": 14, "part_caf": 16,
                "reste": 16, "depot": 16, "date_entree": 13, "date_sortie": 13,
                "caution": 14, "observation": 28, "date_bail": 14, "mode_paiement": 18,
                "jour_echeance": 14}
    for champ, lettre_col in lettre.items():
        ws.column_dimensions[lettre_col].width = largeurs.get(champ, 14)

    for r, loc in enumerate(cfg["locataires"], start=2):
        a_sortie = _date(loc.get("date_sortie")) is not None
        ecrire_texte(ws, r, idx["locataire"], _identite(loc))
        ecrire_texte(ws, r, idx["identifiant"], loc.get("identifiant", ""))
        ecrire_texte(ws, r, idx["adresse"], loc.get("adresse", ""))
        ecrire_texte(ws, r, idx["type_bien"], loc.get("type_bien", TYPES_BIEN[0]))
        if split:
            ws.cell(r, idx["loyer_nu"], _num(loc, "loyer_nu") or 0)
            ws.cell(r, idx["charges"], _num(loc, "charges") or 0)
            ws.cell(r, idx["loyer_total"], f"={lettre['loyer_nu']}{r}+{lettre['charges']}{r}")
        else:
            ws.cell(r, idx["loyer_total"], _num(loc, "loyer", "loyer_total", "loyer_nu") or 0)
        if caf:
            ws.cell(r, idx["part_caf"], _num(loc, "part_caf") or 0)
            ws.cell(r, idx["reste"], f"={lettre['loyer_total']}{r}-{lettre['part_caf']}{r}")
        else:
            ws.cell(r, idx["reste"], f"={lettre['loyer_total']}{r}")
        if depot:
            ws.cell(r, idx["depot"], _num(loc, "depot_garantie") or 0)
        if (de := _date(loc.get("date_entree"))):
            ws.cell(r, idx["date_entree"], de)
        if (ds := _date(loc.get("date_sortie"))):
            ws.cell(r, idx["date_sortie"], ds)
        # Caution rendue / observation : pertinents seulement si le locataire est parti.
        if depot and a_sortie:
            ws.cell(r, idx["caution"], "Oui" if loc.get("caution_rendue") else "Non")
        if a_sortie and loc.get("observation"):
            ecrire_texte(ws, r, idx["observation"], loc.get("observation"))
        if docs:
            db = str(loc.get("date_bail") or "").strip()
            try:   # ISO -> JJ/MM/AAAA pour l'affichage (sinon laissé tel quel)
                db = dt.date.fromisoformat(db).strftime("%d/%m/%Y") if db else ""
            except ValueError:
                pass
            ecrire_texte(ws, r, idx["date_bail"], db)
            ecrire_texte(ws, r, idx["mode_paiement"], loc.get("mode_paiement", ""))
            ecrire_texte(ws, r, idx["jour_echeance"], str(loc.get("jour_echeance", "") or ""))

    derniere = len(cfg["locataires"]) + 1

    fmt_par_champ = {"loyer_nu": FMT_EURO, "charges": FMT_EURO, "loyer_total": FMT_EURO,
                     "part_caf": FMT_EURO, "reste": FMT_EURO, "depot": FMT_EURO,
                     "date_entree": FMT_DATE, "date_sortie": FMT_DATE}
    saisie_champs = {"locataire", "identifiant", "adresse", "type_bien", "loyer_nu", "charges",
                     "part_caf", "depot", "date_entree", "date_sortie", "caution", "observation",
                     "date_bail", "mode_paiement", "jour_echeance"}
    for r in range(2, derniere + 1):
        for champ, i in idx.items():
            cell = ws.cell(r, i)
            est_formule = isinstance(cell.value, str) and cell.value.startswith("=")
            style_cellule(cell, fmt=fmt_par_champ.get(champ), calc=est_formule,
                          saisie=not est_formule and champ in saisie_champs)

    def validation(champ: str, valeurs: list[str], *, bloquant=True) -> None:
        dv = DataValidation(type="list", formula1=_formule_liste(valeurs),
                            allow_blank=True, showErrorMessage=bloquant)
        dv.add(f"{lettre[champ]}2:{lettre[champ]}{derniere}")
        ws.add_data_validation(dv)

    validation("type_bien", TYPES_BIEN)
    if depot:
        validation("caution", ["Oui", "Non"])
    validation("observation", OBSERVATIONS, bloquant=False)  # motif libre autorisé

    der_col = lettre[cols[-1][0]]
    table = Table(displayName="TblLocataires", ref=f"A1:{der_col}{derniere}")
    table.tableStyleInfo = TableStyleInfo(name=CHARTE.table, showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "B2"

    wb.defined_names.add(DefinedName("LocatairesListe",
                                     attr_text=f"Locataires!$A$2:$A${derniere}"))
    wb.defined_names.add(DefinedName("RefLocataires",
                                     attr_text=f"Locataires!$A$2:${der_col}${derniere}"))

    return {"idx": idx, "lettre": lettre, "n": len(cfg["locataires"])}


# --------------------------------------------------------------------------- #
# Une feuille de saisie par locataire
# --------------------------------------------------------------------------- #

def _colonnes_locataire(a_charges: bool, charges_sep: bool, mode: str, caf: bool) -> list[dict]:
    cols = [
        {"key": "locataire", "titre": "Locataire", "w": 20, "kind": "meta", "cache": True},
        {"key": "annee", "titre": "Année", "w": 8, "kind": "meta", "fmt": "0"},
        {"key": "mois", "titre": "Mois", "w": 11, "kind": "meta"},
    ]
    if a_charges:
        # En mode « charges comprises », loyer nu et charges restent calculés mais masqués.
        cols += [
            {"key": "loyer_nu_du", "titre": "Loyer nu dû", "w": 12, "kind": "ref",
             "src": "loyer_nu", "fmt": FMT_EURO, "cache": not charges_sep},
            {"key": "charges_du", "titre": "Charges dues", "w": 12, "kind": "ref",
             "src": "charges", "fmt": FMT_EURO, "cache": not charges_sep},
            {"key": "total_du",
             "titre": "Loyer (charges comprises)" if mode == "comprises" else "Total dû",
             "w": 18 if mode == "comprises" else 12, "kind": "calc", "fmt": FMT_EURO},
        ]
    else:
        cols.append({"key": "total_du", "titre": "Loyer dû", "w": 12, "kind": "ref",
                     "src": "loyer_total", "fmt": FMT_EURO})
    if caf:
        cols += [
            {"key": "caf_attendu", "titre": "CAF attendue", "w": 12, "kind": "ref",
             "src": "part_caf", "fmt": FMT_EURO},
            {"key": "caf_recu", "titre": "CAF reçue", "w": 12, "kind": "input", "fmt": FMT_EURO},
            {"key": "caf_date", "titre": "Date CAF", "w": 12, "kind": "input", "fmt": FMT_DATE},
        ]
    cols += [
        {"key": "rac_attendu", "titre": "Reste à charge attendu", "w": 16, "kind": "calc",
         "fmt": FMT_EURO},
        {"key": "loc_recu", "titre": "Part locataire reçue", "w": 16, "kind": "input",
         "fmt": FMT_EURO},
        {"key": "loc_date", "titre": "Date locataire", "w": 13, "kind": "input", "fmt": FMT_DATE},
        {"key": "total_recu", "titre": "Total reçu", "w": 12, "kind": "calc", "fmt": FMT_EURO},
        {"key": "ecart", "titre": "Écart", "w": 11, "kind": "calc", "fmt": FMT_EURO},
        {"key": "statut", "titre": "Statut", "w": 16, "kind": "calc"},
    ]
    return cols


# Ligne d'en-tête du tableau dans une feuille locataire (après le titre).
PL_LIGNE_ENTETE = 4


def construire_feuilles_locataires(wb: Workbook, cfg: dict, ref_loc: dict,
                                   saisies: dict) -> list[dict]:
    mod = cfg["modules"]
    split, charges_sep, mode = _flags_charges(cfg)
    caf = mod["caf"]
    irl_on = bool(mod.get("irl"))
    lettre_loc = ref_loc["lettre"]
    cols = _colonnes_locataire(split, charges_sep, mode, caf)
    L = {c["key"]: get_column_letter(i) for i, c in enumerate(cols, 1)}
    col_de = {c["key"]: i for i, c in enumerate(cols, 1)}

    pris: set = set()
    infos: list[dict] = []

    # Colonnes totalisées dans la ligne « Total <année> ».
    cols_total = ["total_du"] + (["caf_recu"] if caf else []) + \
        ["loc_recu", "total_recu", "ecart"]

    for loc_index, loc in enumerate(cfg["locataires"]):
        ident_complet = _identite(loc)
        identifiant = str(loc.get("identifiant") or "").strip()
        surname = str(loc.get("nom") or "").strip() or ident_complet
        base = " - ".join(p for p in (identifiant, surname) if p) or ident_complet
        feuille = _nom_feuille(base, pris)
        ws = wb.create_sheet(feuille)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = CHARTE.onglet_locataire
        rloc = loc_index + 2  # ligne du locataire dans l'onglet Locataires

        style_titre(ecrire_texte(ws, 1, 2, ident_complet))
        ecrire_texte(ws, 1, 4, identifiant).font = Font(bold=True, size=TITRE_H2)
        ws.cell(2, 2, "Adresse :").font = Font(bold=True)
        ws.cell(2, 4, "=" + _ref("Locataires", f"${lettre_loc['adresse']}${rloc}"))

        for c in cols:
            i = col_de[c["key"]]
            style_entete(ws.cell(PL_LIGNE_ENTETE, i, c["titre"]))
            ws.column_dimensions[get_column_letter(i)].width = c["w"]
            if c.get("cache"):
                ws.column_dimensions[get_column_letter(i)].hidden = True
        regler_hauteur_entete(ws, PL_LIGNE_ENTETE)

        # rloc est stable sur toute l'itération du locataire ; refloc est appelé en
        # synchrone (pas de capture différée) -> B023 faux positif.
        def refloc(field: str) -> str:
            return "=" + _ref("Locataires", f"${lettre_loc[field]}${rloc}")  # noqa: B023

        # Mois groupés par année (pour insérer un total + une ligne vide entre les années).
        par_annee: dict = {}
        for (annee, m) in _mois_actifs(loc, cfg["annee_debut"], cfg["annee_fin"]):
            par_annee.setdefault(annee, []).append(m)

        rows_map: dict = {}
        r = PL_LIGNE_ENTETE + 1
        for annee in sorted(par_annee):
            y0 = r
            for m in par_annee[annee]:
                nom_mois = MOIS[m - 1]
                preserve = saisies.get((ident_complet, int(annee), nom_mois), {})
                sfx = _prorata_suffixe(loc, annee, m)
                for c in cols:
                    key = c["key"]
                    cell = ws.cell(r, col_de[key])
                    if key == "locataire":
                        cell.value = ident_complet
                        _neutraliser(cell)
                    elif key == "annee":
                        cell.value = annee
                    elif key == "mois":
                        cell.value = nom_mois
                    elif c["kind"] == "ref":
                        # Prorata appliqué au loyer / charges (pas à la CAF, qui se calcule à part).
                        sf = sfx if c["src"] in ("loyer_nu", "charges", "loyer_total") else ""
                        if irl_on and c["src"] in ("loyer_nu", "loyer_total"):
                            # IRL activé : le loyer attendu suit le loyer applicable de l'année
                            # (table « Loyer applicable par année » de l'onglet Révision IRL).
                            cell.value = (f"=SUMIFS(LoyerAn_Valeur,LoyerAn_Loc,${L['locataire']}{r},"
                                          f"LoyerAn_Annee,${L['annee']}{r}){sf}")
                        else:
                            cell.value = refloc(c["src"]) + sf
                    elif key == "total_du" and split:
                        cell.value = f"={L['loyer_nu_du']}{r}+{L['charges_du']}{r}"
                    elif key == "rac_attendu":
                        cell.value = (f"={L['total_du']}{r}-{L['caf_attendu']}{r}" if caf
                                      else f"={L['total_du']}{r}")
                    elif key == "total_recu":
                        cell.value = (f"={L['caf_recu']}{r}+{L['loc_recu']}{r}" if caf
                                      else f"={L['loc_recu']}{r}")
                    elif key == "ecart":
                        cell.value = f"={L['total_recu']}{r}-{L['total_du']}{r}"
                    elif key == "statut":
                        tr, ec = f"{L['total_recu']}{r}", f"{L['ecart']}{r}"
                        td = f"{L['total_du']}{r}"
                        cell.value = (f'=IF(AND({td}=0,{tr}=0),"Soldé",'
                                      f'IF({tr}=0,"À encaisser",'
                                      f'IF(ABS({ec})<=0.005,"Soldé",'
                                      f'IF({ec}>0,"Trop-perçu","Partiel"))))')
                    elif key in COLS_SAISIE and key in preserve:
                        cell.value = preserve[key]
                    style_cellule(cell, fmt=c.get("fmt"),
                                  saisie=c["kind"] == "input",
                                  calc=c["kind"] in ("calc", "ref"))
                rows_map[(annee, m)] = r
                r += 1

            # Ligne « Total <année> » : surlignée dans la teinte du thème, pleine largeur.
            tot = ws.cell(r, col_de["mois"], f"Total {annee}")
            tot.font = Font(bold=True)
            for key in cols_total:
                cc = ws.cell(r, col_de[key], f"=SUM({L[key]}{y0}:{L[key]}{r - 1})")
                cc.number_format = FMT_EURO
                cc.font = Font(bold=True)
                cc.border = BORDURE
            fill_total = PatternFill("solid", fgColor=CHARTE.calc)
            for c in range(1, max(col_de.values()) + 1):
                ws.cell(r, c).fill = fill_total
            r += 2  # total + une ligne vide de séparation

        der = r - 1
        if rows_map:
            premier = PL_LIGNE_ENTETE + 1
            plage_statut = f"{L['statut']}{premier}:{L['statut']}{der}"
            for texte, couleur in (("Soldé", CHARTE.solde), ("Trop-perçu", CHARTE.trop),
                                   ("Partiel", CHARTE.partiel), ("À encaisser", CHARTE.attente)):
                ws.conditional_formatting.add(
                    plage_statut,
                    FormulaRule(formula=[f'${L["statut"]}{premier}="{texte}"'],
                                fill=_fill_cf(couleur)))
        ws.freeze_panes = ws.cell(PL_LIGNE_ENTETE + 1, 4).coordinate

        infos.append({"loc": loc, "nom": ident_complet, "feuille": feuille,
                      "cols": L, "rows": rows_map})

    return infos


# --------------------------------------------------------------------------- #
# Feuille Données (consolidée, masquée) : alimente Bilan + documents
# --------------------------------------------------------------------------- #

def construire_donnees(wb: Workbook, cfg: dict, feuilles: list[dict]) -> None:
    mod = cfg["modules"]
    split, _csep, _mode = _flags_charges(cfg)
    caf = mod["caf"]

    cols = ["locataire", "annee", "mois"]
    if split:
        cols += ["loyer_nu_du", "charges_du"]
    cols.append("total_du")
    if caf:
        cols += ["caf_attendu", "caf_recu"]
    cols += ["loc_recu", "loc_date", "total_recu"]
    pos = {k: i for i, k in enumerate(cols, 1)}

    ws = wb.create_sheet("Données")
    ws.sheet_state = "hidden"
    ws.sheet_properties.tabColor = CHARTE.onglet_donnees
    for i, k in enumerate(cols, 1):
        ws.cell(1, i, k)

    r = 2
    for info in feuilles:
        f, Lpl, rows = info["feuille"], info["cols"], info["rows"]
        for (annee, m), rpl in rows.items():
            ecrire_texte(ws, r, pos["locataire"], info["nom"])
            ws.cell(r, pos["annee"], annee)
            ws.cell(r, pos["mois"], MOIS[m - 1])
            for k in cols[3:]:
                ws.cell(r, pos[k], "=" + _ref(f, f"{Lpl[k]}{rpl}"))
            r += 1
    derniere = max(r - 1, 2)

    def nommer(nom_plage: str, key: str) -> None:
        col = get_column_letter(pos[key])
        wb.defined_names.add(DefinedName(
            nom_plage, attr_text=f"Données!${col}$2:${col}${derniere}"))

    nommer("Suivi_Locataire", "locataire")
    nommer("Suivi_Annee", "annee")
    nommer("Suivi_Mois", "mois")
    nommer("Suivi_TotalDu", "total_du")
    nommer("Suivi_TotalRecu", "total_recu")
    nommer("Suivi_LocRecu", "loc_recu")
    nommer("Suivi_LocDate", "loc_date")
    if split:
        nommer("Suivi_LoyerNuDu", "loyer_nu_du")
        nommer("Suivi_ChargesDu", "charges_du")
    if caf:
        nommer("Suivi_CAFRecue", "caf_recu")


# --------------------------------------------------------------------------- #
# Onglet Bilan
# --------------------------------------------------------------------------- #

def construire_bilan(wb: Workbook, cfg: dict) -> dict:
    """Construit l'onglet Bilan : évolution annuelle (portefeuille) + synthèse par
    locataire (toutes années) + un bloc détail par année. Renvoie les coordonnées
    des blocs (consommées par construire_tableau_bord)."""
    caf = cfg["modules"]["caf"]
    locs = cfg["locataires"]
    annees = list(range(cfg["annee_debut"], cfg["annee_fin"] + 1))
    ws = wb.create_sheet("Bilan")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHARTE.onglet_systeme

    cols = [("Locataire", 24), ("Total dû", 14), ("Total reçu", 14)]
    if caf:
        cols += [("dont CAF", 13), ("dont locataire", 14)]
    cols += [("Solde", 14), ("Taux recouvrement", 16)]
    keys = ["nom", "du", "recu"] + (["caf", "loc"] if caf else []) + ["solde", "taux"]
    B = {k: get_column_letter(i) for i, k in enumerate(keys, 1)}
    pos = {k: i for i, k in enumerate(keys, 1)}
    for i, (_titre, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    fill_total = PatternFill("solid", fgColor=CHARTE.calc)

    def bloc(r0, titre, entete_col1, lignes, cle_col1, extra_crit=""):
        """En-tête + lignes SUMIFS + ligne TOTAL. `lignes` = liste de
        (valeur_col1, est_nombre) ; le critère SUMIFS de chaque ligne référence sa
        propre colonne 1 ($A) via `cle_col1` (la valeur EST le critère), plus
        `extra_crit` (ex. filtre année fixe). Renvoie (hdr, first, last, total)."""
        style_titre(ws.cell(r0, 1, titre), TITRE_H2)
        hdr = r0 + 1
        for i, t in enumerate([entete_col1] + [c[0] for c in cols[1:]], 1):
            style_entete(ws.cell(hdr, i, t))
        regler_hauteur_entete(ws, hdr)
        first = hdr + 1
        r = first
        for valeur, est_nb in lignes:
            if est_nb:
                ws.cell(r, 1, valeur).number_format = "0"
            else:
                ecrire_texte(ws, r, 1, valeur)
            crit = f"{cle_col1},$A{r}{extra_crit}"
            ws.cell(r, pos["du"], f"=SUMIFS(Suivi_TotalDu,{crit})")
            ws.cell(r, pos["recu"], f"=SUMIFS(Suivi_TotalRecu,{crit})")
            if caf:
                ws.cell(r, pos["caf"], f"=SUMIFS(Suivi_CAFRecue,{crit})")
                ws.cell(r, pos["loc"], f"=SUMIFS(Suivi_LocRecu,{crit})")
            ws.cell(r, pos["solde"], f"={B['recu']}{r}-{B['du']}{r}")
            ws.cell(r, pos["taux"], f'=IFERROR({B["recu"]}{r}/{B["du"]}{r},"")')
            r += 1
        last = r - 1
        total = r
        ws.cell(total, 1, "TOTAL").font = Font(bold=True)
        for k in (["du", "recu"] + (["caf", "loc"] if caf else []) + ["solde"]):
            col = B[k]
            ws.cell(total, pos[k], f"=SUM({col}{first}:{col}{last})").font = Font(bold=True)
        ws.cell(total, pos["taux"],
                f'=IFERROR({B["recu"]}{total}/{B["du"]}{total},"")').font = Font(bold=True)
        for rr in range(first, total + 1):
            for k in keys:
                cell = ws.cell(rr, pos[k])
                if k == "taux":
                    cell.number_format = FMT_PCT
                elif k != "nom":
                    cell.number_format = FMT_EURO
                cell.border = BORDURE
                if rr == total:
                    cell.fill = fill_total
        plage = f"{B['solde']}{first}:{B['solde']}{last}"
        ws.conditional_formatting.add(plage, FormulaRule(
            formula=[f"${B['solde']}{first}<-0.005"], fill=_fill_cf(CHARTE.partiel)))
        ws.conditional_formatting.add(plage, FormulaRule(
            formula=[f"${B['solde']}{first}>0.005"], fill=_fill_cf(CHARTE.trop)))
        return hdr, first, last, total

    style_titre(ws.cell(1, 1, "BILAN"))

    # 1) Évolution annuelle (portefeuille) : une ligne par année.
    a_hdr, a_first, a_last, a_total = bloc(
        3, "Évolution annuelle (portefeuille)", "Année",
        [(an, True) for an in annees], "Suivi_Annee")

    # 2) Synthèse par locataire (toutes années) : la table historique.
    g_hdr, g_first, g_last, g_total = bloc(
        a_total + 3, "Synthèse par locataire (toutes années)", "Locataire",
        [(_identite(loc), False) for loc in locs], "Suivi_Locataire")

    # 3) Détail par année : un bloc-table locataire par année de la période.
    r0 = g_total + 3
    for an in annees:
        _h, _f, _l, t = bloc(
            r0, f"Détail {an}", "Locataire",
            [(_identite(loc), False) for loc in locs], "Suivi_Locataire",
            extra_crit=f",Suivi_Annee,{an}")
        r0 = t + 3

    ws.freeze_panes = "A2"
    return {
        "caf": caf, "col": B, "pos": pos,
        "global": {"hdr": g_hdr, "first": g_first, "last": g_last, "total": g_total},
        "annuel": {"hdr": a_hdr, "first": a_first, "last": a_last, "total": a_total},
    }


# --------------------------------------------------------------------------- #
# Onglet Révision IRL (calculateur d'aide)
# --------------------------------------------------------------------------- #

def construire_irl(wb: Workbook, cfg: dict, ref_loc: dict, saisies_irl: dict) -> None:
    if not cfg["modules"].get("irl"):
        return
    a_charges, _csep, _mode = _flags_charges(cfg)
    lettre_loc = ref_loc["lettre"]
    loyer_field = "loyer_nu" if a_charges else "loyer_total"
    annees = list(range(cfg["annee_debut"], cfg["annee_fin"] + 1))

    ws = wb.create_sheet("Révision IRL")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHARTE.onglet_systeme
    for col, w in (("A", 24), ("B", 12), ("C", 14), ("D", 16), ("E", 18), ("F", 16)):
        ws.column_dimensions[col].width = w

    style_titre(ws.cell(1, 1, "RÉVISION DU LOYER (IRL)"))
    ecrire_lien(ws.cell(2, 1), "→ Valeurs officielles de l'IRL (série INSEE)", URL_IRL_INSEE)

    # --- Section 1 : indices IRL (saisie) ---
    style_titre(ws.cell(3, 1, "Indices IRL publiés (INSEE), à saisir"), TITRE_H2)
    for i, t in enumerate(("Année", "Trimestre", "Valeur IRL"), 1):
        style_entete(ws.cell(4, i, t))
    regler_hauteur_entete(ws, 4)
    ws.cell(3, 4, "Saisissez ici les indices ; le loyer révisé de chaque année est "
            "calculé ci-dessous et répercuté dans le suivi mensuel.").font = Font(italic=True)

    idx_saisis = saisies_irl.get("indices", {})
    r = 5
    for annee in annees:
        for t in TRIMESTRES:
            ca = ws.cell(r, 1, annee)
            ca.number_format = "0"
            ca.border = BORDURE
            ws.cell(r, 2, t).border = BORDURE
            cv = ws.cell(r, 3)
            v = idx_saisis.get((int(annee), t))
            if v not in (None, ""):
                cv.value = v
            style_cellule(cv, saisie=True, fmt="0.00")
            r += 1
    fin_idx = r - 1
    sh = "'Révision IRL'"
    wb.defined_names.add(DefinedName("Irl_Annee", attr_text=f"{sh}!$A$5:$A${fin_idx}"))
    wb.defined_names.add(DefinedName("Irl_Trim", attr_text=f"{sh}!$B$5:$B${fin_idx}"))
    wb.defined_names.add(DefinedName("Irl_Valeur", attr_text=f"{sh}!$C$5:$C${fin_idx}"))

    # --- Section 2 : loyer applicable par année (révision IRL répercutée) ---
    # Modèle fermé : loyer(année Y) = loyer_base × IRL_Tref(Y) / IRL_Tref(A0), où A0 est la
    # première année de présence (loyer de base = fiche Locataires) et Tref le trimestre de
    # référence dérivé de la date d'entrée. Indices absents → IFERROR retombe sur le loyer de
    # base (aucune révision). Ces valeurs alimentent le suivi mensuel via les plages LoyerAn_*.
    rs = fin_idx + 3
    style_titre(ws.cell(rs - 1, 1, "Loyer applicable par année (révision répercutée)"), TITRE_H2)
    entetes = ["Locataire", "Année", "Trimestre réf.", "Loyer de base (€)",
               "Loyer applicable (€)", "Variation vs base"]
    for i, t in enumerate(entetes, 1):
        style_entete(ws.cell(rs, i, t))
    regler_hauteur_entete(ws, rs)

    r = rs + 1
    for loc_index, loc in enumerate(cfg["locataires"]):
        nom = _identite(loc)
        rloc = loc_index + 2
        base_ref = _ref("Locataires", f"${lettre_loc[loyer_field]}${rloc}")
        annees_loc = _annees_actives(loc, cfg["annee_debut"], cfg["annee_fin"])
        if not annees_loc:
            continue
        a0 = annees_loc[0]                       # année de référence (loyer de base)
        trim = _trimestre_de(loc.get("date_entree"))
        for annee in annees_loc:
            ecrire_texte(ws, r, 1, nom).border = BORDURE
            ca = ws.cell(r, 2, annee)
            ca.number_format = "0"
            ca.border = BORDURE
            ws.cell(r, 3, trim).border = BORDURE
            cb = ws.cell(r, 4, "=" + base_ref)
            cb.number_format = FMT_EURO
            cb.border = BORDURE
            # Numérateur = indice de l'année courante ; dénominateur = indice de
            # l'année de référence (a0). Un indice non publié rend son SUMIFS nul :
            # on retombe alors sur le loyer de base (pas de révision), au lieu de
            # laisser 0/x=0 écraser le loyer. IFERROR garde la division par zéro
            # (réf. absente) ; le IF garde le numérateur absent (année non publiée).
            num = f'SUMIFS(Irl_Valeur,Irl_Annee,$B{r},Irl_Trim,$C{r})'
            den = f'SUMIFS(Irl_Valeur,Irl_Annee,{a0},Irl_Trim,$C{r})'
            capp = ws.cell(r, 5,
                           f'=IF({num}=0,$D{r},IFERROR($D{r}*{num}/{den},$D{r}))')
            capp.number_format = FMT_EURO
            capp.border = BORDURE
            cvar = ws.cell(r, 6, f'=IFERROR($E{r}/$D{r}-1,"")')
            cvar.number_format = FMT_PCT
            cvar.border = BORDURE
            r += 1
    fin_loy = r - 1

    if fin_loy >= rs + 1:
        sh2 = "'Révision IRL'"
        wb.defined_names.add(DefinedName(
            "LoyerAn_Loc", attr_text=f"{sh2}!$A${rs + 1}:$A${fin_loy}"))
        wb.defined_names.add(DefinedName(
            "LoyerAn_Annee", attr_text=f"{sh2}!$B${rs + 1}:$B${fin_loy}"))
        wb.defined_names.add(DefinedName(
            "LoyerAn_Valeur", attr_text=f"{sh2}!$E${rs + 1}:$E${fin_loy}"))
    ws.freeze_panes = "A5"


# --------------------------------------------------------------------------- #
# Onglet Tableau de bord (graphiques)
# --------------------------------------------------------------------------- #

def construire_tableau_bord(wb: Workbook, cfg: dict, layout: dict) -> None:
    if not cfg["modules"].get("tableau_bord", True) or "Bilan" not in wb.sheetnames:
        return
    if len(cfg["locataires"]) == 0:
        return
    caf = layout["caf"]
    B, pos = layout["col"], layout["pos"]
    glob, ann = layout["global"], layout["annuel"]
    bilan = wb["Bilan"]

    ws = wb.create_sheet("Tableau de bord")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHARTE.onglet_systeme
    style_titre(ws.cell(1, 1, "TABLEAU DE BORD"))
    ws.cell(2, 1, "Synthèse visuelle, calculée à partir de vos saisies dans les feuilles "
            "locataire. Les chiffres se mettent à jour automatiquement.").font = Font(
        italic=True, color=CHARTE.onglet_donnees)

    # --- Cartes de synthèse (lues sur la ligne TOTAL de la synthèse globale) ---
    # Colonnes des cartes élargies : un total comme « 123 456 € » en gros doit tenir
    # sans déborder en « ### ». Sans incidence sur les graphes (largeur fixée en cm).
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 12
    fmt_eur_kpi = '#,##0\\ "€"'           # KPI : pas de centimes (lisibilité)
    gt = glob["total"]
    kpis = [("Total dû", f"='Bilan'!{B['du']}{gt}", fmt_eur_kpi),
            ("Total reçu", f"='Bilan'!{B['recu']}{gt}", fmt_eur_kpi),
            ("Solde (reçu − dû)", f"='Bilan'!{B['solde']}{gt}", fmt_eur_kpi),
            ("Taux de recouvrement", f"='Bilan'!{B['taux']}{gt}", FMT_PCT)]
    fill_kpi = PatternFill("solid", fgColor=CHARTE.calc)
    for i, (lab, frm, fmt) in enumerate(kpis):
        c0 = 2 + i * 3                      # cartes ancrées en B, E, H, K (2 col. chacune)
        for rr in (4, 5):
            ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c0 + 1)
            for cc in range(c0, c0 + 2):
                cell = ws.cell(rr, cc)
                cell.fill = fill_kpi
                cell.border = BORDURE
        lc = ws.cell(4, c0, lab)
        lc.font = Font(bold=True, size=10, color=CHARTE.onglet_donnees)
        vc = ws.cell(5, c0, frm)
        vc.font = Font(bold=True, size=16, color=CHARTE.primaire)
        vc.number_format = fmt
    ws.row_dimensions[5].height = 26

    def graphe(g, ancre_row, caption):
        cap = ws.cell(ancre_row - 1, 2, caption)
        cap.font = Font(italic=True, color=CHARTE.onglet_donnees)
        g.height, g.width = 8, 16
        g.legend.position = "b"
        # Style intégré 2 = « Style 1 » de la galerie Excel (barres pleines colorées,
        # coins arrondis) : rendu soigné dès la génération, sans repasser par Excel.
        g.style = 2
        g.roundedCorners = True
        ws.add_chart(g, f"B{ancre_row}")

    cats_loc = Reference(bilan, min_col=pos["nom"], min_row=glob["first"], max_row=glob["last"])

    g1 = BarChart()
    g1.type = "col"
    g1.title = "Loyers : dû vs reçu par locataire"
    g1.y_axis.title = "€"
    d1 = Reference(bilan, min_col=pos["du"], max_col=pos["recu"],
                   min_row=glob["hdr"], max_row=glob["last"])
    g1.add_data(d1, titles_from_data=True)
    g1.set_categories(cats_loc)
    graphe(g1, 8, "Pour chaque locataire : montant attendu (dû) face au montant réellement "
                  "encaissé, sur toute la période.")

    gA = BarChart()
    gA.type = "col"
    gA.title = "Évolution annuelle : dû vs reçu"
    gA.y_axis.title = "€"
    dA = Reference(bilan, min_col=pos["du"], max_col=pos["recu"],
                   min_row=ann["hdr"], max_row=ann["last"])
    gA.add_data(dA, titles_from_data=True)
    gA.set_categories(Reference(bilan, min_col=1, min_row=ann["first"], max_row=ann["last"]))
    graphe(gA, 26, "Total dû et total encaissé, année par année (tous locataires confondus). "
                   "Un écart croissant signale des impayés qui s'accumulent.")

    g2 = BarChart()
    g2.type = "bar"
    g2.title = "Taux de recouvrement par locataire"
    g2.x_axis.title = "% encaissé"
    d2 = Reference(bilan, min_col=pos["taux"], min_row=glob["hdr"], max_row=glob["last"])
    g2.add_data(d2, titles_from_data=True)
    g2.set_categories(cats_loc)
    graphe(g2, 44, "Part des loyers encaissée par locataire. 100 % = tout payé ; en dessous, "
                   "il reste des sommes à percevoir.")

    if caf:
        g3 = BarChart()
        g3.type = "col"
        g3.grouping = "stacked"
        g3.overlap = 100
        g3.title = "Répartition de l'encaissé : CAF / locataire"
        g3.y_axis.title = "€"
        d3 = Reference(bilan, min_col=pos["caf"], max_col=pos["loc"],
                       min_row=glob["hdr"], max_row=glob["last"])
        g3.add_data(d3, titles_from_data=True)
        g3.set_categories(cats_loc)
        graphe(g3, 62, "Origine de l'argent reçu : part versée par la CAF (tiers payant) "
                       "et part payée directement par le locataire.")


# --------------------------------------------------------------------------- #
# Documents à imprimer (quittance, avis d'échéance, lettre de relance)
# --------------------------------------------------------------------------- #

def construire_document(wb: Workbook, cfg: dict, ref_loc: dict, kind: str) -> None:
    mod = cfg["modules"]
    split, _csep, _mode = _flags_charges(cfg)
    caf = mod["caf"]
    idx = ref_loc["idx"]
    annees = list(range(cfg["annee_debut"], cfg["annee_fin"] + 1))

    # base : « recu » = atteste un paiement (quittance) ; « du » = appelle/réclame.
    specs = {
        "quittance": {
            "feuille": "Quittance", "titre": "QUITTANCE DE LOYER",
            "base": "recu",
            "note": "Quittance valable uniquement si le loyer est intégralement réglé ; "
                    "à défaut, le présent document vaut reçu de paiement partiel (art. 21 "
                    "de la loi du 6 juillet 1989). Envoi par voie électronique soumis à "
                    "l'accord exprès du locataire.",
        },
        "avis": {
            "feuille": "Avis d'échéance", "titre": "AVIS D'ÉCHÉANCE",
            "base": "du",
            "note": "Ce document ne peut tenir lieu de quittance.",
        },
        "relance": {
            "feuille": "Lettre de relance", "titre": "LETTRE DE RELANCE",
            "base": "arrieres",
            "note": "Rappel amiable. À défaut de régularisation, il pourra être suivi d'une "
                    "mise en demeure par lettre recommandée avec accusé de réception.",
        },
        "mise_en_demeure": {
            "feuille": "Mise en demeure", "titre": "MISE EN DEMEURE DE PAYER",
            "base": "arrieres",
            "note": "À adresser par lettre recommandée avec accusé de réception. Ce courrier "
                    "ne vaut pas commandement de payer (acte de commissaire de justice).",
        },
    }
    spec = specs[kind]

    b = cfg["bailleur"]
    iban = str(b.get("iban") or "").strip()   # seul champ document au niveau bailleur
    # mode_paiement / jour_echeance / date_bail sont propres au locataire : lus par
    # VLOOKUP sur le locataire selectionne (voir vl_mode / vl_jour / vl_bail plus bas).
    any_mode = any(str(loc.get("mode_paiement") or "").strip()
                   for loc in cfg["locataires"])

    ws = wb.create_sheet(spec["feuille"])
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHARTE.onglet_document
    for col, w in (("A", 3), ("B", 26), ("C", 26), ("D", 16), ("E", 16)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:E2")
    titre = ws["B2"]
    titre.value = spec["titre"]   # quittance : remplacé par une formule conditionnelle plus bas
    titre.font = Font(bold=True, size=18, color=CHARTE.primaire)
    titre.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26
    # Filet d'accent du thème sous le titre, sur toute la largeur du contenu.
    for col in range(2, 6):
        ws.cell(2, col).border = Border(bottom=Side(style="medium", color=CHARTE.primaire))

    # Sélecteurs.
    ws["B4"] = "Locataire :"
    ecrire_texte(ws, 4, 3, _identite(cfg["locataires"][0]))
    ws["B5"], ws["C5"] = "Mois :", MOIS[0]
    ws["D5"], ws["E5"] = "Année :", annees[0]
    for lab in ("B4", "B5", "D5"):
        ws[lab].font = Font(bold=True)
    for sel in ("C4", "C5", "E5"):
        ws[sel].fill = PatternFill("solid", fgColor=CHARTE.saisie)
        ws[sel].border = BORDURE

    def valider(cellule: str, formule: str) -> None:
        dv = DataValidation(type="list", formula1=formule, allow_blank=False)
        dv.add(cellule)
        ws.add_data_validation(dv)

    valider("C4", "=LocatairesListe")
    valider("C5", _formule_liste(MOIS))
    valider("E5", _formule_liste(annees))

    def vlook(field: str) -> str:
        return f'IFERROR(VLOOKUP($C$4,RefLocataires,{idx[field]},FALSE),"")'

    # Champs document propres au locataire (présents dans le référentiel quand le
    # module documents est actif) : lus dynamiquement selon le locataire sélectionné.
    vl_mode = vlook("mode_paiement") if "mode_paiement" in idx else '""'
    vl_jour = vlook("jour_echeance") if "jour_echeance" in idx else '""'
    vl_bail = vlook("date_bail") if "date_bail" in idx else '""'

    # Bloc bailleur (gauche) et locataire (droite).
    ws["B7"] = "Le bailleur :"
    ws["B7"].font = Font(bold=True, color=CHARTE.primaire)
    ecrire_texte(ws, 8, 2, b.get("nom", ""))
    for i, k in enumerate(("adresse", "tel", "email"), start=9):
        if b.get(k):
            ecrire_texte(ws, i, 2, str(b[k]))

    ws["D7"] = "Le locataire :"
    ws["D7"].font = Font(bold=True, color=CHARTE.primaire)
    ws["D8"] = "=$C$4"
    ws["D9"] = f"={vlook('identifiant')}"
    ws["D10"] = f"={vlook('adresse')}"

    crit = "Suivi_Locataire,$C$4,Suivi_Annee,$E$5,Suivi_Mois,$C$5"

    def sif(plage: str) -> str:
        return f"SUMIFS({plage},{crit})"

    # Reste dû cumulé (toutes périodes) pour le locataire sélectionné : critère sur
    # le seul locataire (relance / mise en demeure réclament la dette globale, pas
    # le seul mois sélectionné). Pas de détail mois par mois (modèle à sélecteur unique).
    def sif_loc(plage: str) -> str:
        return f"SUMIFS({plage},Suivi_Locataire,$C$4)"

    # ROUND au centime : Suivi_TotalDu est calculé en direct (prorata => fractions de
    # centime) alors que Suivi_TotalRecu vient d'une saisie arrondie ; sans arrondi, un
    # mois soldé apparaîtrait « restant dû 0,001 € » et fausserait la bascule quittance.
    reste_global = f'ROUND({sif_loc("Suivi_TotalDu")}-{sif_loc("Suivi_TotalRecu")},2)'

    # Tableau des montants. Chaque ligne porte une clé (ou None) repérant les totaux
    # réutilisés par le titre/corps, pour ne pas dépendre de la position des lignes.
    r0 = 13
    lignes = []   # (label, formule, cle)
    if spec["base"] == "recu":   # quittance
        if split:
            lignes += [("Loyer nu", sif("Suivi_LoyerNuDu"), None),
                       ("Charges", sif("Suivi_ChargesDu"), None)]
        lignes.append(("Total loyer + charges dû", sif("Suivi_TotalDu"), "du"))
        if caf:
            lignes += [("Dont part CAF reçue", sif("Suivi_CAFRecue"), None),
                       ("Dont part locataire reçue", sif("Suivi_LocRecu"), None)]
        lignes.append(("Montant total reçu", sif("Suivi_TotalRecu"), "recu"))
    elif spec["base"] == "arrieres":   # relance / mise en demeure
        if split:
            lignes += [("Loyer pour la période", sif("Suivi_LoyerNuDu"), None),
                       ("Charges (provisions)", sif("Suivi_ChargesDu"), None)]
        lignes.append(("Total dû (période)", sif("Suivi_TotalDu"), "du"))
        lignes.append(("Déjà reçu (période)", sif("Suivi_TotalRecu"), "recu"))
        lignes.append(("Reste dû (période)",
                       f'ROUND({sif("Suivi_TotalDu")}-{sif("Suivi_TotalRecu")},2)', "reste"))
        lignes.append(("Reste dû — toutes périodes", reste_global, "reste_global"))
    else:   # avis
        if split:
            lignes += [("Loyer pour la période", sif("Suivi_LoyerNuDu"), None),
                       ("Charges (provisions)", sif("Suivi_ChargesDu"), None)]
        lignes.append(("Total à régler", sif("Suivi_TotalDu"), "du"))

    # Récap encadré : libellé (B) + valeur (C). Les lignes de total sont surlignées
    # dans la teinte douce du thème pour ressortir sans alourdir.
    _fill_total = PatternFill("solid", fgColor=CHARTE.calc)

    def _ligne_recap(r: int, label: str, *, gras: bool, total: bool):
        lab = ws.cell(r, 2, label)
        lab.font = Font(bold=gras)
        lab.alignment = Alignment(horizontal="left", vertical="center")
        lab.border = BORDURE
        val = ws.cell(r, 3)
        val.font = Font(bold=gras)
        val.alignment = Alignment(horizontal="right", vertical="center")
        val.border = BORDURE
        if total:
            lab.fill = val.fill = _fill_total
        return val

    cellule = {}   # cle -> référence de cellule "$C$<row>"
    for i, (label, formule, cle) in enumerate(lignes):
        r = r0 + i
        total = cle is not None
        val = _ligne_recap(r, label, gras=total, total=total)
        val.value = f"={formule}"
        val.number_format = FMT_EURO
        if cle:
            cellule[cle] = f"$C${r}"

    # Titre conditionnel de la quittance : reçu intégral => quittance, sinon reçu partiel.
    if kind == "quittance":
        du, recu = cellule["du"], cellule["recu"]
        # Soldé au centime près => quittance ; sinon reçu partiel (cf. ROUND ci-dessus).
        titre.value = (f'=IF(AND({du}>0,ROUND({du}-{recu},2)<=0),"QUITTANCE DE LOYER",'
                       f'"REÇU DE PAIEMENT PARTIEL")')

    fin = r0 + len(lignes)
    if kind == "quittance":
        r_date = fin
        cd = _ligne_recap(r_date, "Date de paiement", gras=True, total=False)
        cd.value = f'=IF({sif("Suivi_LocRecu")}=0,"",{sif("Suivi_LocDate")})'
        cd.number_format = FMT_DATE
        fin = r_date + 1

    # Corps (curseur de ligne : le corps, puis des blocs optionnels, puis la signature).
    if kind == "quittance":
        du, recu = cellule["du"], cellule["recu"]
        plein = (f'"Je soussigné(e) "&$B$8&", bailleur, déclare avoir reçu de "&$C$4'
                 f'&" la somme de "&FIXED({recu},2,TRUE)&" € au titre du loyer et des charges '
                 f'pour la période de "&$C$5&" "&$E$5'
                 f'&", et lui en donne quittance, sous réserve de tous mes droits."')
        partiel = (f'"Je soussigné(e) "&$B$8&", bailleur, déclare avoir reçu de "&$C$4'
                   f'&" la somme de "&FIXED({recu},2,TRUE)&" € à valoir sur le loyer et les '
                   f'charges de la période de "&$C$5&" "&$E$5&" (montant dû : "'
                   f'&FIXED({du},2,TRUE)&" €, restant dû : "&FIXED(ROUND({du}-{recu},2),2,TRUE)'
                   f'&" €). Le présent reçu ne vaut pas quittance."')
        corps = f"=IF(AND({du}>0,ROUND({du}-{recu},2)<=0),{plein},{partiel})"
        corps_rows = 4
    elif kind == "avis":
        montant_cell = cellule["du"]
        base_txt = (f'"Madame, Monsieur, veuillez trouver le montant de votre loyer pour la '
                    f'période de "&$C$5&" "&$E$5&", soit "&FIXED({montant_cell},2,TRUE)&" €"')
        # Echeance selon le jour du bail du locataire (vide => delai par defaut).
        ech = (f'&IF({vl_jour}="",", à régler sous 8 jours.",'
               f'", à régler avant le "&{vl_jour}&" "&$C$5&" "&$E$5&".")')
        corps = "=" + base_txt + ech
        corps_rows = 3
    elif kind == "relance":
        corps = (f'="Madame, Monsieur, sauf erreur de notre part, le loyer de la période de "'
                 f'&$C$5&" "&$E$5&" demeure impayé. À ce jour, le solde restant dû pour ce '
                 f'logement est de "&FIXED({cellule["reste_global"]},2,TRUE)&" €. Nous vous '
                 f'remercions de régulariser cette somme dans les meilleurs délais."')
        corps_rows = 4
    else:   # mise_en_demeure
        # Référence au bail du locataire (vide => phrase omise) via IF dans la formule.
        corps = (f'="Madame, Monsieur, "&IF({vl_bail}="","",'
                 f'"En exécution du bail conclu le "&{vl_bail}&", ")'
                 f'&"nous vous mettons en demeure de régler sous 8 jours la somme de "'
                 f'&FIXED({cellule["reste_global"]},2,TRUE)&" € correspondant aux loyers et '
                 f'charges restant dus pour le logement situé "&$D$10&". À défaut de règlement '
                 f'dans ce délai, nous nous réservons le droit de faire délivrer un commandement '
                 f'de payer par commissaire de justice visant la clause résolutoire du bail, '
                 f'puis de saisir le juge des contentieux de la protection."')
        corps_rows = 5

    r = fin + 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r + corps_rows - 1, end_column=5)
    ws.cell(r, 2, corps).alignment = Alignment(wrap_text=True, vertical="top")
    r += corps_rows + 1

    # Modalités de paiement (avis + mise en demeure). IBAN = bailleur (statique) ;
    # mode de paiement = locataire (VLOOKUP selon le locataire sélectionné).
    if kind in ("avis", "mise_en_demeure") and (iban or any_mode):
        ws.cell(r, 2, "Modalités de paiement :").font = Font(bold=True, color=CHARTE.primaire)
        r += 1
        if any_mode:
            ws.cell(r, 2, "Mode :").font = Font(bold=True)
            ws.cell(r, 3, f"={vl_mode}")
            r += 1
        if iban:
            ws.cell(r, 2, "IBAN :").font = Font(bold=True)
            ecrire_texte(ws, r, 3, iban)
            r += 1
        r += 1

    ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=5)
    note = ws.cell(r, 2, spec["note"])
    note.font = Font(italic=True)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    r += 4
    ws.cell(r, 2, "Fait à ……………………………, le ……………………………")
    r += 2
    ws.cell(r, 2, "Signature du bailleur :").font = Font(bold=True)

    # Document destiné à l'impression : une page A4 portrait, centrée.
    mettre_en_page_impression(ws, f"A1:E{r}")


def construire_documents(wb: Workbook, cfg: dict, ref_loc: dict) -> None:
    for kind in ("quittance", "avis", "relance", "mise_en_demeure"):
        construire_document(wb, cfg, ref_loc, kind)


# --------------------------------------------------------------------------- #
# Onglet Régularisation des charges (annuelle, par locataire)
# --------------------------------------------------------------------------- #

def construire_regularisation(wb: Workbook, cfg: dict, saisies_reg: dict) -> None:
    # Sans charges, il n'y a pas de provisions à régulariser.
    a_charges, _csep, mode = _flags_charges(cfg)
    if not (cfg["modules"].get("regularisation_charges") and a_charges):
        return

    ws = wb.create_sheet("Régularisation charges")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHARTE.onglet_systeme
    titres = [("Locataire", 24), ("Année", 10), ("Provisions appelées (€)", 20),
              ("Charges réelles (€)", 18), ("Solde (€)", 14), ("Sens", 28)]
    for i, (t, w) in enumerate(titres, 1):
        style_entete(ws.cell(1, i, t))
        ws.column_dimensions[get_column_letter(i)].width = w
    regler_hauteur_entete(ws, 1)

    r = 2
    for loc in cfg["locataires"]:
        nom = _identite(loc)
        annees = sorted({a for (a, _) in _mois_actifs(loc, cfg["annee_debut"], cfg["annee_fin"])})
        for annee in annees:
            ecrire_texte(ws, r, 1, nom).border = BORDURE
            ca = ws.cell(r, 2, annee)
            ca.number_format = "0"
            ca.border = BORDURE
            cp = ws.cell(r, 3, f"=SUMIFS(Suivi_ChargesDu,Suivi_Locataire,$A{r},Suivi_Annee,$B{r})")
            cp.number_format = FMT_EURO
            cp.border = BORDURE
            cr = ws.cell(r, 4)
            v = saisies_reg.get((str(nom), int(annee)))
            if v not in (None, ""):
                cr.value = v
            elif mode == "comprises":
                # Charges comprises : par défaut, charges réelles = provisions (modifiable).
                cr.value = f"=C{r}"
            style_cellule(cr, saisie=True, fmt=FMT_EURO)
            cs = ws.cell(r, 5, f"=C{r}-D{r}")
            cs.number_format = FMT_EURO
            cs.border = BORDURE
            csens = ws.cell(r, 6, f'=IF(D{r}=0,"Charges réelles à saisir",'
                                  f'IF(ABS(C{r}-D{r})<=0.005,"Équilibré",'
                                  f'IF(C{r}>D{r},"À rembourser au locataire","Complément à demander")))')
            csens.border = BORDURE
            r += 1
    der = r - 1
    if der >= 2:
        ws.conditional_formatting.add(f"E2:E{der}", FormulaRule(
            formula=["$E2<-0.005"], fill=_fill_cf(CHARTE.partiel)))
        ws.conditional_formatting.add(f"E2:E{der}", FormulaRule(
            formula=["$E2>0.005"], fill=_fill_cf(CHARTE.trop)))
        ws.auto_filter.ref = f"A1:F{der}"  # filtre par locataire / année
    ws.freeze_panes = "A2"


# --------------------------------------------------------------------------- #
# Onglet Guide
# --------------------------------------------------------------------------- #

def construire_guide(wb: Workbook, cfg: dict) -> None:
    ws = wb.create_sheet("Guide", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHARTE.onglet_systeme
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 86

    blanc_centre = Alignment(horizontal="left", vertical="center")
    texte_wrap = Alignment(wrap_text=True, vertical="center")

    def fusion(row, valeur=None):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        return ws.cell(row, 2, valeur)

    def section(row, titre):
        """Barre de section : titre blanc sur bandeau bleu (cohérent avec l'en-tête)."""
        c = fusion(row, titre)
        c.font = Font(bold=True, size=TITRE_H2, color=CHARTE.entete_txt)
        c.fill = PatternFill("solid", fgColor=CHARTE.primaire)
        c.alignment = blanc_centre
        ws.row_dimensions[row].height = 22
        return row + 1

    def pastille(row, etiquette, couleur, desc, *, texte_blanc=False):
        """Ligne de légende : pastille colorée en B + description en C."""
        chip = ws.cell(row, 2, etiquette)
        chip.fill = PatternFill("solid", fgColor=couleur)
        chip.alignment = Alignment(horizontal="center", vertical="center")
        chip.border = BORDURE
        if texte_blanc:
            chip.font = Font(bold=True, color=CHARTE.entete_txt)
        d = ws.cell(row, 3, desc)
        d.alignment = texte_wrap
        return row + 1

    # --- Bandeau d'en-tête (bailleur) ---
    b = cfg["bailleur"]
    coord = " · ".join(str(b[k]) for k in ("adresse", "tel", "email") if b.get(k))
    for rr in range(1, 4):
        for cc in range(1, 4):
            ws.cell(rr, cc).fill = PatternFill("solid", fgColor=CHARTE.primaire)
    titre = fusion(1, "Suivi des loyers")
    titre.font = Font(bold=True, size=20, color=CHARTE.entete_txt)
    titre.alignment = blanc_centre
    ws.row_dimensions[1].height = 30
    sous = fusion(2, f"Bailleur : {b.get('nom', '')}")
    sous.font = Font(bold=True, size=TITRE_H2, color=CHARTE.entete_txt)
    sous.alignment = blanc_centre
    ws.row_dimensions[2].height = 18
    cco = fusion(3)
    cco.alignment = blanc_centre
    if coord:
        _neutraliser(ws.cell(3, 2, coord))
    cco.font = Font(italic=True, color=CHARTE.entete_txt)
    ws.row_dimensions[3].height = 16

    # --- Comment ça marche (stepper) ---
    r = section(5, "Comment ça marche")
    etapes = [
        "Onglet « Locataires » : vérifiez les biens (type, n° / nom, adresse), les loyers de "
        "référence et les dates d'entrée / sortie. Les mois d'entrée et de sortie partiels sont "
        "calculés au prorata des jours.",
        "Une feuille par locataire : chaque mois, saisissez les montants REÇUS dans les cellules "
        "jaunes (CAF reçue, part locataire reçue, dates).",
        "Totaux, écarts et statuts se calculent automatiquement (cellules bleutées, à ne pas "
        "modifier).",
        "Onglet « Bilan » : synthèse par locataire (total dû, reçu, solde, taux de recouvrement).",
    ]
    if cfg["modules"].get("tableau_bord", True):
        etapes.append("Onglet « Tableau de bord » (2e onglet) : graphiques dû vs reçu et taux "
                      "de recouvrement, mis à jour automatiquement.")
    if cfg["modules"].get("documents"):
        etapes.append("Documents à imprimer (Quittance, Avis d'échéance, Lettre de relance) : "
                      "choisissez le locataire et la période dans les listes déroulantes ; le "
                      "document se remplit seul.")
    if cfg["modules"].get("regularisation_charges") and _flags_charges(cfg)[0]:
        etapes.append("Onglet « Régularisation charges » : saisissez les charges réelles "
                      "annuelles (pré-remplies en mode charges comprises) ; le solde par "
                      "locataire se calcule seul.")
    if cfg["modules"].get("irl"):
        etapes.append("Onglet « Révision IRL » : saisissez les indices IRL publiés ; le loyer "
                      "révisé de chaque année est calculé et répercuté automatiquement dans le "
                      "loyer attendu du suivi mensuel (lien officiel ci-dessous).")
    for i, texte in enumerate(etapes, 1):
        pastille(r, str(i), CHARTE.onglet_systeme, texte, texte_blanc=True)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1

    # --- Code couleur des onglets (rappelle la charte des onglets) ---
    r = section(r, "Repère des onglets")
    r = pastille(r, "", CHARTE.onglet_systeme,
                 "Bleu : pilotage et synthèse (Guide, Tableau de bord, Locataires, Bilan…).")
    r = pastille(r, "", CHARTE.onglet_locataire,
                 "Vert : une feuille de saisie par locataire (c'est là qu'on remplit chaque mois).")
    if cfg["modules"].get("documents"):
        r = pastille(r, "", CHARTE.onglet_document,
                     "Orange : documents à imprimer (quittance, avis d'échéance, relance).")
    r += 1

    # --- Légende des statuts ---
    r = section(r, "Légende des statuts")
    for nom, couleur, desc in (
        ("Soldé", CHARTE.solde, "Le total reçu couvre le total dû."),
        ("Trop-perçu", CHARTE.trop, "Reçu supérieur au dû (avance ou régularisation à prévoir)."),
        ("Partiel", CHARTE.partiel, "Reçu inférieur au dû (impayé partiel)."),
        ("À encaisser", CHARTE.attente, "Aucun paiement saisi pour ce mois."),
    ):
        r = pastille(r, nom, couleur, desc)

    # --- Liens utiles ---
    if cfg["modules"].get("irl"):
        r += 1
        r = section(r, "Liens utiles")
        fusion(r)
        ecrire_lien(ws.cell(r, 2), "Valeurs officielles de l'IRL (série trimestrielle INSEE)",
                    URL_IRL_INSEE)


# --------------------------------------------------------------------------- #
# Préservation des saisies (lecture des feuilles locataire)
# --------------------------------------------------------------------------- #

_TITRE_VERS_KEY = {
    "CAF reçue": "caf_recu", "Date CAF": "caf_date",
    "Part locataire reçue": "loc_recu", "Date locataire": "loc_date",
}


def recolter_saisies(wb) -> dict:
    """Saisies utilisateur d'un classeur déjà chargé : (nom, année, mois) -> {colonne: valeur}."""
    saisies: dict = {}

    for nom_feuille in wb.sheetnames:
        if nom_feuille in FEUILLES_SYSTEME:
            continue
        ws = wb[nom_feuille]
        # Repérer la ligne d'en-tête (celle qui contient « Mois » ET « Année »).
        ligne_ent = None
        for r in range(1, min(ws.max_row, 12) + 1):
            valeurs = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if "Mois" in valeurs and "Année" in valeurs:
                ligne_ent = r
                entetes = {v: c for c, v in enumerate(valeurs, 1) if v}
                break
        if ligne_ent is None:
            continue
        c_loc, c_an, c_mo = (entetes.get("Locataire"), entetes.get("Année"), entetes.get("Mois"))
        if not all((c_loc, c_an, c_mo)):
            continue
        for r in range(ligne_ent + 1, ws.max_row + 1):
            nom = ws.cell(r, c_loc).value
            annee = ws.cell(r, c_an).value
            mois = ws.cell(r, c_mo).value
            if not nom or annee in (None, "") or not mois:
                continue
            valeurs = {}
            for titre, key in _TITRE_VERS_KEY.items():
                col = entetes.get(titre)
                if col:
                    v = ws.cell(r, col).value
                    if v not in (None, ""):
                        valeurs[key] = v
            if valeurs:
                saisies[(str(nom), int(annee), str(mois))] = valeurs
    return saisies


def recolter_regularisation(wb) -> dict:
    """Charges réelles déjà saisies : (nom, année) -> montant."""
    if "Régularisation charges" not in wb.sheetnames:
        return {}
    ws = wb["Régularisation charges"]
    ent = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    cL, cA, cR = ent.get("Locataire"), ent.get("Année"), ent.get("Charges réelles (€)")
    if not all((cL, cA, cR)):
        return {}
    res: dict = {}
    for r in range(2, ws.max_row + 1):
        nom, an, v = ws.cell(r, cL).value, ws.cell(r, cA).value, ws.cell(r, cR).value
        if nom and an not in (None, "") and v not in (None, ""):
            res[(str(nom), int(an))] = v
    return res


def recolter_irl(wb) -> dict:
    """Indices IRL déjà saisis : {"indices": {(année, trimestre): valeur}}.

    Seuls les indices sont saisis ; la table « Loyer applicable par année » est entièrement
    recalculée à la régénération, donc rien d'autre à préserver.
    """
    if "Révision IRL" not in wb.sheetnames:
        return {}
    ws = wb["Révision IRL"]
    res: dict = {"indices": {}}

    h_idx = None
    for r in range(1, ws.max_row + 1):
        ligne = [ws.cell(r, c).value for c in range(1, 7)]
        if "Valeur IRL" in ligne:
            h_idx = r
            break

    if h_idx:
        r = h_idx + 1
        while r <= ws.max_row:
            an, tr, v = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
            if an in (None, "") or tr in (None, ""):
                break
            if v not in (None, ""):
                res["indices"][(int(an), str(tr))] = v
            r += 1
    return res


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def generer_workbook(cfg: dict, sortie: Path, *, preserver: bool = True,
                     orphelins_out: list | None = None) -> Path:
    cfg = valider_config(cfg) if "annee_debut" not in cfg else cfg
    sortie = Path(sortie)

    # Identité graphique : résout la charte active pour toute la génération.
    global CHARTE
    CHARTE = resoudre_charte(cfg.get("theme"), cfg.get("police"))

    # Un seul chargement du classeur existant pour récolter toutes les saisies.
    saisies, saisies_reg, saisies_irl = {}, {}, {}
    if preserver and sortie.is_file():
        wbx = load_workbook(sortie, data_only=False)
        saisies = recolter_saisies(wbx)
        saisies_reg = recolter_regularisation(wbx)
        saisies_irl = recolter_irl(wbx)

    # Exemples : pré-remplissage de démonstration si aucune saisie réelle à préserver.
    if cfg.get("demo") and not saisies:
        saisies = _saisies_demo(cfg)

    # Signaler les saisies orphelines (locataire renommé/supprimé) : non réinjectées.
    identites = {_identite(loc) for loc in cfg["locataires"]}
    orphelins = sorted({nom for (nom, _, _) in saisies if nom not in identites})
    if orphelins:
        print("Attention : saisies non réattribuées (locataire renommé ou supprimé) : "
              + ", ".join(orphelins), file=sys.stderr)
        if orphelins_out is not None:
            orphelins_out.extend(orphelins)

    wb = Workbook()
    wb.remove(wb.active)

    ref_loc = construire_locataires(wb, cfg)
    feuilles = construire_feuilles_locataires(wb, cfg, ref_loc, saisies)
    construire_donnees(wb, cfg, feuilles)
    layout_bilan = construire_bilan(wb, cfg)
    construire_tableau_bord(wb, cfg, layout_bilan)
    construire_regularisation(wb, cfg, saisies_reg)
    construire_irl(wb, cfg, ref_loc, saisies_irl)
    if cfg["modules"].get("documents"):
        construire_documents(wb, cfg, ref_loc)
    construire_guide(wb, cfg)

    # Tableau de bord en 2e position, juste après le Guide (visible dès l'ouverture).
    if "Tableau de bord" in wb.sheetnames:
        wb.move_sheet("Tableau de bord", offset=1 - wb.sheetnames.index("Tableau de bord"))

    # Impression des onglets larges : paysage, ajusté en largeur (hauteur libre), pour
    # rester lisible sans modifier les largeurs de colonnes. Les documents gardent leur
    # mise en page portrait une page (posée dans construire_document).
    for nom in ("Bilan", "Régularisation charges", "Révision IRL"):
        if nom in wb.sheetnames:
            ws = wb[nom]
            mettre_en_page_impression(ws, ws.dimensions, paysage=True,
                                      hauteur_pages=0, centre=False)
    # Le Guide est une fiche de référence : on le force sur une seule page (largeur ET
    # hauteur), pour que « Liens utiles » ne déborde jamais en page 2.
    if "Guide" in wb.sheetnames:
        guide = wb["Guide"]
        mettre_en_page_impression(guide, guide.dimensions, paysage=True,
                                  hauteur_pages=1, centre=False)
    if "Tableau de bord" in wb.sheetnames:
        # Graphiques hors plage de cellules : pas de zone explicite, le tableur les inclut.
        mettre_en_page_impression(wb["Tableau de bord"], None, paysage=True,
                                  hauteur_pages=0, centre=False)

    # Police d'identité, en dernière passe (couvre aussi les cellules sans style explicite),
    # puis compensation de largeur des colonnes pour les polices plus larges que Calibri.
    appliquer_police(wb, CHARTE.police)
    ajuster_colonnes(wb, CHARTE.police)

    sortie.parent.mkdir(parents=True, exist_ok=True)
    # Sauvegarde de secours avant d'écraser (récupération en cas de couac).
    # Nom horodaté : on conserve chaque sauvegarde au lieu d'écraser la précédente.
    if sortie.is_file():
        try:
            horo = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
            shutil.copy2(sortie, sortie.with_name(f"{sortie.stem}.bak-{horo}.xlsx"))
        except OSError:
            pass
    wb.save(sortie)
    return sortie


def generer(chemin_config: Path, dossier_sortie: Path) -> Path:
    cfg = charger_config(chemin_config)
    sortie = Path(dossier_sortie) / f"Suivi_{base_slug(cfg['bailleur'])}.xlsx"
    return generer_workbook(cfg, sortie)


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 2
    chemin_config = Path(argv[1])
    dossier_sortie = Path(argv[2]) if len(argv) > 2 else Path("sorties")
    if not chemin_config.is_file():
        print(f"Config introuvable : {chemin_config}", file=sys.stderr)
        return 1
    print(f"✔ Classeur généré : {generer(chemin_config, dossier_sortie)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
