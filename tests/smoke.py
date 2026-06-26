"""Smoke test du moteur, autonome (dossier temporaire, ne pollue pas sorties/).

Couvre : architecture (feuille par locataire + Données masquée + documents), modularité,
période d'activité (rotation), et préservation des saisies lors d'une régénération.
"""

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

import generer_suivi_loyers as g

CFG_COMPLET = {
    "bailleur": {"nom": "Smoke Complet"},
    "periode": {"annee_debut": 2024, "annee_fin": 2025},
    "modules": {"mode_charges": "separees", "caf": True, "depot_garantie": True,
                "documents": True, "tableau_bord": True, "regularisation_charges": True,
                "irl": True},
    "locataires": [
        # Nom + prénom, parti en cours de période (caution + observation).
        {"nom": "Alice", "prenom": "A", "type_bien": "Appartement", "identifiant": "Appt 1",
         "adresse": "1 rue A", "loyer_nu": 500, "charges": 50, "part_caf": 200,
         "depot_garantie": 500, "date_entree": "2024-01-01", "date_sortie": "2024-06-30",
         "caution_rendue": False, "observation": "Travaux"},
        {"nom": "Bob", "type_bien": "Appartement", "identifiant": "Appt 2",
         "adresse": "1 rue A", "loyer_nu": 400, "charges": 40, "part_caf": 0,
         "depot_garantie": 400, "date_entree": "2025-07-01"},  # entré en cours de période
    ],
}

CFG_MINIMAL = {
    "bailleur": {"nom": "Smoke Minimal"},
    "periode": {"annee_debut": 2025, "annee_fin": 2025},
    "modules": {"loyer_nu_charges": False, "caf": False, "depot_garantie": False,
                "documents": False},
    "locataires": [{"nom": "Carl", "type_bien": "Maison", "identifiant": "Maison",
                    "adresse": "2 rue B", "loyer": 750, "date_entree": "2025-01-01"}],
}


def ligne_entete(ws):
    for r in range(1, 12):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "Mois" in vals and "Année" in vals:
            return r, {v: c for c, v in enumerate(vals, 1) if v}
    raise AssertionError("en-tête introuvable")


def main() -> int:
    tmp = Path(tempfile.mkdtemp())

    # 1) Bailleur complet : architecture attendue.
    f1 = tmp / "complet.xlsx"
    g.generer_workbook(g.valider_config(CFG_COMPLET), f1)
    wb = load_workbook(f1)
    # Onglet locataire nommé « identifiant - Nom » (évite les doublons même appartement).
    # Tableau de bord en 2e position (juste après le Guide).
    attendus = ["Guide", "Tableau de bord", "Locataires", "Appt 1 - Alice", "Appt 2 - Bob",
                "Données", "Bilan", "Régularisation charges", "Révision IRL",
                "Quittance", "Avis d'échéance", "Lettre de relance"]
    assert wb.sheetnames == attendus, wb.sheetnames
    assert wb["Données"].sheet_state == "hidden"
    assert wb["Quittance"]["B2"].value == "QUITTANCE DE LOYER"
    assert wb["Avis d'échéance"]["B2"].value == "AVIS D'ÉCHÉANCE"

    # Charte : couleur d'onglet par rôle (bleu système / vert locataire / orange document /
    # gris Données) et gridlines masquées partout. Suit la charte résolue du moteur
    # (thème par défaut = « classique »).
    ch = g.resoudre_charte()
    def couleur_onglet(nom):
        tc = wb[nom].sheet_properties.tabColor
        return tc.rgb[-6:] if tc and tc.rgb else None
    role = {"Guide": ch.onglet_systeme, "Locataires": ch.onglet_systeme, "Bilan": ch.onglet_systeme,
            "Tableau de bord": ch.onglet_systeme, "Régularisation charges": ch.onglet_systeme,
            "Révision IRL": ch.onglet_systeme, "Données": ch.onglet_donnees,
            "Appt 1 - Alice": ch.onglet_locataire, "Appt 2 - Bob": ch.onglet_locataire,
            "Quittance": ch.onglet_document, "Avis d'échéance": ch.onglet_document,
            "Lettre de relance": ch.onglet_document}
    for nom, attendu in role.items():
        assert couleur_onglet(nom) == attendu, (nom, couleur_onglet(nom), attendu)
    for nom in wb.sheetnames:
        if wb[nom].sheet_state == "visible":
            assert wb[nom].sheet_view.showGridLines is False, nom

    # Guide redessiné : bandeau de titre + lien INSEE en pied (module IRL actif).
    assert wb["Guide"]["B1"].value == "Suivi des loyers", wb["Guide"]["B1"].value
    irl = wb["Révision IRL"]
    liens = [irl.cell(r, c).hyperlink.target for r in range(1, 5)
             for c in range(1, 4) if irl.cell(r, c).hyperlink]
    assert g.URL_IRL_INSEE in liens, liens

    # Identité = « Nom Prénom ».
    assert g._identite({"nom": "Martin", "prenom": "Sophie"}) == "Martin Sophie"
    assert g._identite({"nom": "Old"}) == "Old"

    # Nom de fichier : SCI prioritaire si renseignée.
    assert g.base_fichier({"nom": "Durand", "prenom": "Paul"}) == "Durand Paul"
    assert g.base_fichier({"nom": "Durand", "sci": True, "sci_nom": "SCI Lilas"}) == "SCI Lilas Durand"
    assert g.base_fichier({"nom": "Durand", "sci": True, "sci_nom": ""}) == "Durand"

    # Sécurité : un champ utilisateur commençant par « = » ne doit pas devenir une formule.
    inj = {
        "bailleur": {"nom": "=cmd|' /C calc'!A1"},
        "periode": {"annee_debut": 2025, "annee_fin": 2025},
        "modules": {"loyer_nu_charges": False, "caf": False, "depot_garantie": False,
                    "documents": True},
        "locataires": [{"nom": "=HYPERLINK(1)", "identifiant": "=2+2", "adresse": "=A1",
                        "loyer": 100, "date_entree": "2025-01-01"}],
    }
    fi = tmp / "inj.xlsx"
    g.generer_workbook(g.valider_config(inj), fi)
    wbi = load_workbook(fi)
    cli = wbi["Locataires"].cell(2, 1)
    assert cli.data_type == "s" and cli.value == "=HYPERLINK(1)", (cli.data_type, cli.value)
    cdoc = wbi["Quittance"].cell(4, 3)  # sélecteur locataire du document
    assert cdoc.data_type == "s", cdoc.data_type

    # Validations (M1/M2/m8).
    bse = {"bailleur": {"nom": "B"}, "periode": {"annee_debut": 2025, "annee_fin": 2025},
           "modules": {"loyer_nu_charges": False, "caf": False, "depot_garantie": False,
                       "documents": False}}

    def doit_lever(cfg, motif):
        try:
            g.valider_config(cfg)
        except ValueError:
            return
        raise AssertionError(f"ValueError attendue : {motif}")

    doit_lever({**bse, "locataires": [
        {"nom": "Dupont", "prenom": "Jean", "loyer": 100, "date_entree": "2025-01-01"},
        {"nom": "Dupont", "prenom": "Jean", "loyer": 100, "date_entree": "2025-01-01"}]},
        "doublon d'identité")
    doit_lever({**bse, "locataires": [
        {"nom": "X", "loyer": "abc", "date_entree": "2025-01-01"}]}, "montant non numérique")
    doit_lever({**bse, "locataires": [
        {"nom": "Y", "loyer": 100, "date_entree": "2025-06-01", "date_sortie": "2025-01-01"}]},
        "sortie avant entrée")

    # m8 : casse normalisée + module inconnu averti.
    _, av = g.migrer_config({**bse, "modules": {"Documents": True, "bidon": True},
                             "locataires": [{"nom": "Z", "loyer": 100, "date_entree": "2025-01-01"}]})
    assert any("inconnu" in a.lower() for a in av), av
    norm = g.valider_config({**bse, "modules": {"Documents": True},
                             "locataires": [{"nom": "Z", "loyer": 100, "date_entree": "2025-01-01"}]})
    assert norm["modules"]["documents"] is True

    # Bloc 1 : version plus récente, coercition virgule, .bak + orphelins.
    _, avv = g.migrer_config({**bse, "version": 999,
                              "locataires": [{"nom": "Z", "loyer": 100, "date_entree": "2025-01-01"}]})
    assert any("récente" in a.lower() for a in avv), avv
    g.valider_config({**bse, "locataires": [
        {"nom": "V", "loyer": "100,5", "date_entree": "2025-01-01"}]})  # ne doit pas lever

    fb = tmp / "bak.xlsx"
    g.generer_workbook(g.valider_config({**bse, "locataires": [
        {"nom": "Ana", "identifiant": "M1", "loyer": 100, "date_entree": "2025-01-01"}]}), fb)
    wbb = load_workbook(fb)
    sa = wbb["M1 - Ana"]
    re_, ea = ligne_entete(sa)
    sa.cell(re_ + 1, ea["Part locataire reçue"]).value = 77
    wbb.save(fb)
    orph: list = []
    g.generer_workbook(g.valider_config({**bse, "locataires": [
        {"nom": "Bea", "identifiant": "M1", "loyer": 100, "date_entree": "2025-01-01"}]}),
        fb, orphelins_out=orph)
    assert "Ana" in orph, orph
    assert (tmp / "bak.bak.xlsx").is_file(), "sauvegarde .bak manquante"

    # Prorata : entrée le 6 janvier (31 jours) -> facteur *26/31 dans le loyer dû.
    fp = tmp / "prorata.xlsx"
    g.generer_workbook(g.valider_config({**bse, "locataires": [
        {"nom": "Pro", "identifiant": "P1", "loyer": 1000, "date_entree": "2025-01-06"}]}), fp)
    wp = load_workbook(fp)["P1 - Pro"]
    rp, ep = ligne_entete(wp)
    assert "*26/31" in str(wp.cell(rp + 1, ep["Loyer dû"]).value), wp.cell(rp + 1, ep["Loyer dû"]).value

    # Mode charges comprises : colonne combinée + charges réelles pré-remplies (=C2).
    fc = tmp / "comprises.xlsx"
    g.generer_workbook(g.valider_config({
        "bailleur": {"nom": "C"}, "periode": {"annee_debut": 2025, "annee_fin": 2025},
        "modules": {"mode_charges": "comprises", "caf": False, "depot_garantie": False,
                    "documents": False, "regularisation_charges": True},
        "locataires": [{"nom": "Cc", "identifiant": "K1", "loyer_nu": 400, "charges": 50,
                        "date_entree": "2025-01-01"}]}), fc)
    wc = load_workbook(fc)
    _, eh = ligne_entete(wc["K1 - Cc"])
    assert "Loyer (charges comprises)" in eh, list(eh)
    assert str(wc["Régularisation charges"].cell(2, 4).value) == "=C2", \
        wc["Régularisation charges"].cell(2, 4).value

    # Onglet Locataires : Type après Adresse + colonnes Caution / Observation renseignées.
    loc = wb["Locataires"]
    hl = {loc.cell(1, c).value: c for c in range(1, loc.max_column + 1)}
    assert list(hl).index("Type de bien") > list(hl).index("Adresse du logement")
    assert loc.cell(2, hl["Caution rendue"]).value == "Non"
    assert loc.cell(2, hl["Observation (motif de départ)"]).value == "Travaux"

    # Feuille locataire : titre = identité, colonnes CAF présentes.
    ws = wb["Appt 1 - Alice"]
    assert ws.cell(1, 2).value == "Alice A"
    assert ws.cell(1, 4).value == "Appt 1"
    _, ent = ligne_entete(ws)
    assert "CAF reçue" in ent and "Charges dues" in ent, list(ent)

    # 2) Période d'activité : Bob entré 07/2025 -> 6 mois. Totaux annuels présents.
    wsb = wb["Appt 2 - Bob"]
    rb, _ = ligne_entete(wsb)
    nb_b = sum(1 for r in range(rb + 1, wsb.max_row + 1)
               if wsb.cell(r, 3).value in g.MOIS)
    assert nb_b == 6, f"Bob devrait avoir 6 mois, obtenu {nb_b}"
    assert any(str(wsb.cell(r, 3).value or "").startswith("Total")
               for r in range(rb + 1, wsb.max_row + 1)), "ligne Total annuel manquante"

    # 3) Bailleur minimal : pas de CAF, pas de documents.
    f2 = tmp / "minimal.xlsx"
    g.generer_workbook(g.valider_config(CFG_MINIMAL), f2)
    wb2 = load_workbook(f2)
    assert "Quittance" not in wb2.sheetnames, wb2.sheetnames
    assert "Maison - Carl" in wb2.sheetnames, wb2.sheetnames
    _, ent2 = ligne_entete(wb2["Maison - Carl"])
    assert not any("CAF" in (k or "") for k in ent2), list(ent2)
    assert "Loyer dû" in ent2, list(ent2)

    # 4) Préservation : on saisit dans la feuille d'Alice, on régénère avec un locataire en plus.
    wb = load_workbook(f1)
    ws = wb["Appt 1 - Alice"]
    r_ent, ent = ligne_entete(ws)
    r0 = r_ent + 1
    ws.cell(r0, ent["Part locataire reçue"]).value = 333
    cle = (ws.cell(r0, ent["Locataire"]).value, ws.cell(r0, ent["Année"]).value,
           ws.cell(r0, ent["Mois"]).value)
    wb.save(f1)

    cfg2 = g.valider_config(CFG_COMPLET)
    cfg2["locataires"].append({"nom": "Dora", "type_bien": "Maison", "identifiant": "Villa",
                               "adresse": "9 rue C", "loyer_nu": 300, "charges": 20,
                               "part_caf": 100, "depot_garantie": 300,
                               "date_entree": "2025-01-01"})
    g.generer_workbook(cfg2, f1, preserver=True)

    wb = load_workbook(f1)
    assert "Villa - Dora" in wb.sheetnames, wb.sheetnames
    ws = wb["Appt 1 - Alice"]
    r_ent, ent = ligne_entete(ws)
    trouve = None
    for r in range(r_ent + 1, ws.max_row + 1):
        if (ws.cell(r, ent["Locataire"]).value, ws.cell(r, ent["Année"]).value,
                ws.cell(r, ent["Mois"]).value) == cle:
            trouve = ws.cell(r, ent["Part locataire reçue"]).value
            break
    assert trouve == 333, f"saisie non préservée: {trouve}"

    # 5) Régularisation : saisie d'une charge réelle, préservée après régénération.
    wb = load_workbook(f1)
    reg = wb["Régularisation charges"]
    ent = {reg.cell(1, c).value: c for c in range(1, reg.max_column + 1)}
    reg.cell(2, ent["Charges réelles (€)"]).value = 612
    cle_reg = (reg.cell(2, ent["Locataire"]).value, reg.cell(2, ent["Année"]).value)
    wb.save(f1)
    g.generer_workbook(g.valider_config(CFG_COMPLET), f1, preserver=True)
    reg = load_workbook(f1)["Régularisation charges"]
    ent = {reg.cell(1, c).value: c for c in range(1, reg.max_column + 1)}
    val = None
    for r in range(2, reg.max_row + 1):
        if (reg.cell(r, ent["Locataire"]).value, reg.cell(r, ent["Année"]).value) == cle_reg:
            val = reg.cell(r, ent["Charges réelles (€)"]).value
            break
    assert val == 612, f"charge réelle non préservée: {val}"

    # 5b) IRL : saisie d'un indice et d'un choix de révision, préservés après régénération.
    wb = load_workbook(f1)
    irl = wb["Révision IRL"]
    # 1er indice (ligne 5, colonne C) et 1re révision locataire.
    irl.cell(5, 3).value = 145.17                      # valeur IRL T1 année début
    h_rev = next(r for r in range(1, irl.max_row + 1)
                 if irl.cell(r, 7).value == "Nouveau loyer (€)")
    irl.cell(h_rev + 1, 3).value = "T2"                # trimestre réf. 1er locataire
    nom_irl = irl.cell(h_rev + 1, 1).value
    wb.save(f1)
    g.generer_workbook(g.valider_config(CFG_COMPLET), f1, preserver=True)
    irl = load_workbook(f1)["Révision IRL"]
    assert irl.cell(5, 3).value == 145.17, irl.cell(5, 3).value
    h_rev = next(r for r in range(1, irl.max_row + 1)
                 if irl.cell(r, 7).value == "Nouveau loyer (€)")
    assert irl.cell(h_rev + 1, 1).value == nom_irl
    assert irl.cell(h_rev + 1, 3).value == "T2", irl.cell(h_rev + 1, 3).value

    # 6) Rétro-compatibilité : une « ancienne » config (champ bien, module quittances)
    #    se migre et génère sans erreur.
    legacy = {
        "bailleur": {"nom": "Legacy"},
        "periode": {"annee_debut": 2025, "annee_fin": 2025},
        "modules": {"loyer_nu_charges": True, "caf": True, "depot_garantie": False,
                    "quittances": True},
        "locataires": [{"nom": "Old", "bien": "Appt Z", "loyer_nu": 400, "charges": 30,
                        "part_caf": 100, "date_entree": "2025-01-01"}],
    }
    migre, avertis = g.migrer_config(legacy)
    assert migre["modules"].get("documents") is True, migre["modules"]
    assert migre["locataires"][0]["identifiant"] == "Appt Z"
    assert avertis, "des avertissements de migration étaient attendus"
    f3 = tmp / "legacy.xlsx"
    g.generer_workbook(g.valider_config(legacy), f3)
    wb3 = load_workbook(f3)
    assert "Appt Z - Old" in wb3.sheetnames and "Quittance" in wb3.sheetnames, wb3.sheetnames

    # 7) Pré-remplissage de démonstration : demo=true remplit les loyers reçus (mois <= cutoff),
    #    laisse vide après, et reste inerte sans le drapeau.
    demo_cfg = {"bailleur": {"nom": "Demo"}, "demo": True,
                "periode": {"annee_debut": 2025, "annee_fin": 2026},
                "modules": {"loyer_nu_charges": False, "caf": False, "depot_garantie": False,
                            "documents": False},
                "locataires": [{"nom": "Demo", "identifiant": "D1", "loyer": 600,
                                "date_entree": "2025-01-01"}]}
    fd = tmp / "demo.xlsx"
    g.generer_workbook(g.valider_config(demo_cfg), fd, preserver=False)
    wd = load_workbook(fd)["D1 - Demo"]
    rd, ed = ligne_entete(wd)
    recus = {}
    for row in range(rd + 1, wd.max_row + 1):
        an, mo = wd.cell(row, ed["Année"]).value, wd.cell(row, ed["Mois"]).value
        if mo in g.MOIS:
            recus[(an, mo)] = wd.cell(row, ed["Part locataire reçue"]).value
    assert recus.get((2025, "Janvier")) == 600, recus.get((2025, "Janvier"))  # mois plein soldé
    assert recus.get((2026, "Juin")) in (None, ""), recus.get((2026, "Juin"))  # après cutoff = vide

    # Sans le drapeau demo : aucune saisie pré-remplie.
    g.generer_workbook(g.valider_config({**demo_cfg, "demo": False}), fd, preserver=False)
    wn = load_workbook(fd)["D1 - Demo"]
    rn, en = ligne_entete(wn)
    assert wn.cell(rn + 1, en["Part locataire reçue"]).value in (None, ""), "demo off devrait être vide"

    # 8) Thèmes sélectionnables : chaque thème pose sa couleur primaire sur les onglets
    #    système et applique la police d'identité ; thème inconnu -> défaut + avertissement.
    assert g.valider_config(CFG_COMPLET)["theme"] == g.THEME_DEFAUT == "classique"
    assert g.valider_config(CFG_COMPLET)["police"] == g.POLICE_DEFAUT == "Tahoma"
    for tid, spec in g.THEMES.items():
        ct = dict(CFG_COMPLET, theme=tid)
        ft = tmp / f"theme_{tid}.xlsx"
        g.generer_workbook(g.valider_config(ct), ft, preserver=False)
        wt = load_workbook(ft)
        tc = wt["Bilan"].sheet_properties.tabColor
        assert tc and tc.rgb[-6:] == spec["primaire"], (tid, tc.rgb if tc else None)
        assert wt["Guide"]["B1"].font.name == "Tahoma", (tid, wt["Guide"]["B1"].font.name)
    cfg_inc, av_inc = g.migrer_config(dict(CFG_COMPLET, theme="nimportequoi"))
    assert cfg_inc["theme"] == "classique" and any("inconnu" in a for a in av_inc), (cfg_inc["theme"], av_inc)
    # Police personnalisée respectée.
    fpol = tmp / "police.xlsx"
    g.generer_workbook(g.valider_config(dict(CFG_COMPLET, police="Verdana")), fpol, preserver=False)
    assert load_workbook(fpol)["Guide"]["B1"].font.name == "Verdana"

    print("SMOKE OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
